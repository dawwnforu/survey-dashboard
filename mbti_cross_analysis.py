# -*- coding: utf-8 -*-
"""MBTI理论映射 + Q43人格 × 冲突行为 交叉分析"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from scipy import stats as sp_stats
from collections import Counter

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

def c(col):
    return [ws.cell(row, col).value for row in range(2, ws.max_row + 1)]

# Q43人格 (col 76-91, 16题)
# L/O作息: Q43_1,5,9,13  (cols 76,80,84,88) - 每题1-2分, 总分0-4
# T/M秩序: Q43_2,6,10,14 (cols 77,81,85,89)
# D/Z冲突: Q43_3,7,11,15 (cols 78,82,86,90)
# E/I社交: Q43_4,8,12,16 (cols 79,83,87,91)

def calc_q43_dim(cols):
    """Calculate Q43 dimension score (0-4), higher=second letter"""
    scores = []
    for i in range(N):
        row = i + 2
        vals = [ws.cell(row, col).value for col in cols]
        valid = [v for v in vals if v is not None and isinstance(v, (int,float))]
        if len(valid) >= 2:
            scores.append(sum(1 for v in valid if v == 2))  # count "type 2" answers (0-4)
        else:
            scores.append(None)
    return scores

# Q43 dimensions (higher = second letter: O夜猫, M随性, Z佛系, I社恐)
L_O = calc_q43_dim([76,80,84,88])  # 0=晨鸟L, 4=夜猫O
T_M = calc_q43_dim([77,81,85,89])  # 0=整洁T, 4=随性M
D_Z = calc_q43_dim([78,82,86,90])  # 0=直球D, 4=佛系Z
E_I = calc_q43_dim([79,83,87,91])  # 0=社牛E, 4=社恐I

# Conflict resolution (col 30): 1=直接沟通, 2=第三方调解, 4=忍耐
resolve = c(30)
anxiety = c(23)
relation = c(31)
disturbed = c(28)
conflict = c(25)
switch = c(20)
schedule = c(33)
hygiene = c(32)
fairness = c(44)

print("=" * 80)
print("MBTI × 宿舍行为 理论映射与交叉分析")
print("=" * 80)

# === Part 0: 理论映射框架 ===
print("\n" + "=" * 80)
print("一、MBTI ↔ Q43 ↔ 大五人格 ↔ 宿舍行为 四层映射")
print("=" * 80)
print("""
┌──────────┬─────────────┬──────────────┬──────────────────────────┐
│ MBTI维度  │ Q43维度      │ 大五对应      │ 宿舍行为预测              │
├──────────┼─────────────┼──────────────┼──────────────────────────┤
│ E/I 社交  │ E/I 社交能量 │ 外向性(E)     │ I型→需要独处→易被吵醒     │
│          │              │              │ E型→社交缓冲→关系修复快    │
├──────────┼─────────────┼──────────────┼──────────────────────────┤
│ S/N 感知  │ (未直接测量)  │ 开放性(O)     │ S型→关注具体(卫生/噪音)    │
│          │              │              │ N型→关注氛围(关系/公平)    │
├──────────┼─────────────┼──────────────┼──────────────────────────┤
│ T/F 决策  │ D/Z 冲突模式 │ 宜人性(A)     │ T型→直接沟通→冲突多但解决快 │
│          │              │              │ F型→忍耐/调解→内化焦虑     │
├──────────┼─────────────┼──────────────┼──────────────────────────┤
│ J/P 生活  │ T/M 空间秩序 │ 尽责性(C)     │ J型→要求整洁/规律→易不满   │
│          │ L/O 作息节律 │              │ P型→随性→被J型抱怨         │
└──────────┴─────────────┴──────────────┴──────────────────────────┘

INFP典型画像: I(社恐) + N(抽象关注氛围) + F(和谐优先) + P(随性)
  → Q43对应: E/I=社恐 + D/Z=佛系(不直接冲突) + T/M=随性(不太挑剔)
  → 预测行为: 敏感于宿舍氛围, 不满时倾向于"忍耐"而非"直接沟通"
  → 风险: 情绪内耗→高焦虑, 长期压抑→突然爆发或调宿

ESTJ典型画像: E(社牛) + S(关注具体) + T(逻辑优先) + J(结构秩序)
  → Q43对应: E/I=社牛 + D/Z=直球 + T/M=整洁(高要求)
  → 预测行为: 直接指出室友问题, 要求建立宿舍规则
  → 风险: 被认为"事多/强势", 引发人际关系紧张
""")

# === Part 1: Q43人格 → 冲突解决方式 ===
print("=" * 80)
print("二、Q43人格维度 × 冲突解决方式 (验证MBTI预测)")
print("=" * 80)

for dim_name, dim_scores, dim_label in [
    ('D/Z 冲突模式(直球→佛系)', D_Z, '直球D型(0-1分) vs 佛系Z型(3-4分)'),
    ('E/I 社交能量(社牛→社恐)', E_I, '社牛E型(0-1分) vs 社恐I型(3-4分)'),
    ('T/M 空间秩序(整洁→随性)', T_M, '整洁T型(0-1分) vs 随性M型(3-4分)'),
    ('L/O 作息节律(晨鸟→夜猫)', L_O, '晨鸟L型(0-1分) vs 夜猫O型(3-4分)'),
]:
    print(f"\n▶ {dim_name}")

    # Split into low (0-1) vs high (3-4) on each dimension
    low_idx = [i for i in range(N) if dim_scores[i] is not None and dim_scores[i] <= 1]
    high_idx = [i for i in range(N) if dim_scores[i] is not None and dim_scores[i] >= 3]

    if len(low_idx) < 5 or len(high_idx) < 5:
        print(f"  样本不足: low={len(low_idx)}, high={len(high_idx)}")
        continue

    for group_name, idxs in [(dim_label.split('vs')[0].strip(), low_idx),
                              (dim_label.split('vs')[1].strip(), high_idx)]:
        res = [resolve[i] for i in idxs if isinstance(resolve[i], (int,float))]
        cnt = Counter(res)
        total = len(res)
        direct = cnt.get(1, 0)
        third = cnt.get(2, 0)
        endure = cnt.get(4, 0)
        print(f"  {group_name} (N={total}): 直接沟通{direct}({direct/total*100:.0f}%) "
              f"第三方调解{third}({third/total*100:.0f}%) "
              f"忍耐{endure}({endure/total*100:.0f}%)")

    # Chi-square test
    lr = [resolve[i] for i in low_idx if isinstance(resolve[i], (int,float))]
    hr = [resolve[i] for i in high_idx if isinstance(resolve[i], (int,float))]

    # Direct+Third vs Endure (沟通 vs 忍耐)
    l_comm = sum(1 for r in lr if r in [1,2])
    l_endure = sum(1 for r in lr if r == 4)
    h_comm = sum(1 for r in hr if r in [1,2])
    h_endure = sum(1 for r in hr if r == 4)

    if min(l_comm, l_endure, h_comm, h_endure) >= 3:
        table = [[l_comm, l_endure], [h_comm, h_endure]]
        chi2, p, dof, exp = sp_stats.chi2_contingency(table)
        sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
        print(f"  卡方(沟通vs忍耐): chi2={chi2:.2f}, p={p:.4f}{sig}")

# === Part 2: "INFP型"特征 × 宿舍适应 ===
print("\n" + "=" * 80)
print('三、模拟"INFP型"(高I+高F/佛系) vs "ESTJ型"(高E+高T/直球) 宿舍适应差异')
print("=" * 80)

# INFP-like: E_I >= 3 (I) AND D_Z >= 3 (F/佛系)
# ESTJ-like: E_I <= 1 (E) AND D_Z <= 1 (T/直球) AND T_M <= 1 (J/整洁)
# Also need reasonable sample

infp_like = [i for i in range(N) if E_I[i] is not None and D_Z[i] is not None
             and E_I[i] >= 3 and D_Z[i] >= 3]
estj_like = [i for i in range(N) if E_I[i] is not None and D_Z[i] is not None and T_M[i] is not None
             and E_I[i] <= 1 and D_Z[i] <= 1 and T_M[i] <= 1]

print(f"INFP型(I高+佛系): N={len(infp_like)}")
print(f"ESTJ型(E高+直球+整洁): N={len(estj_like)}")

if len(infp_like) >= 5 and len(estj_like) >= 5:
    for name, col, unit in [('焦虑天数', 23, '天/周'), ('被吵醒次数', 28, '次/周'),
                             ('舍友关系', 31, '1-5'), ('冲突次数', 25, '次/月'),
                             ('调宿念头', 20, '1-4'), ('作息一致性', 33, '1-5'),
                             ('公平感', 44, '1公平3不公')]:
        infp_vals = [ws.cell(i+2, col).value for i in infp_like
                     if isinstance(ws.cell(i+2, col).value, (int,float))]
        estj_vals = [ws.cell(i+2, col).value for i in estj_like
                     if isinstance(ws.cell(i+2, col).value, (int,float))]
        if len(infp_vals) >= 3 and len(estj_vals) >= 3:
            m1, m2 = np.mean(infp_vals), np.mean(estj_vals)
            pooled = np.sqrt((np.std(infp_vals, ddof=1)**2 + np.std(estj_vals, ddof=1)**2) / 2)
            d = (m1 - m2) / pooled if pooled > 0 else 0
            t, p = sp_stats.ttest_ind(infp_vals, estj_vals)
            sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
            direction = 'INFP>' if m1 > m2 else 'ESTJ>'
            print(f"  {name}: INFP{m1:.2f} vs ESTJ{m2:.2f}, d={d:+.3f}, p={p:.4f}{sig} {direction}")

# === Part 3: 冲突解决方式 × 各人格维度均值 ===
print("\n" + "=" * 80)
print("四、三种冲突解决方式的人格画像")
print("=" * 80)

for res_type, res_val, res_name in [(1, '直接沟通'), (2, '第三方调解'), (4, '忍耐')]:
    idxs = [i for i in range(N) if isinstance(resolve[i], (int,float)) and resolve[i] == res_val]
    print(f"\n▶ {res_name} (N={len(idxs)})")
    for dim_name, dim_scores, low_label, high_label in [
        ('E/I社交', E_I, '社牛E', '社恐I'),
        ('D/Z冲突', D_Z, '直球D', '佛系Z'),
        ('T/M秩序', T_M, '整洁T', '随性M'),
        ('L/O作息', L_O, '晨鸟L', '夜猫O'),
    ]:
        vals = [dim_scores[i] for i in idxs if dim_scores[i] is not None]
        if vals:
            m = np.mean(vals)
            # 0-4 scale, higher=second letter
            direction = high_label if m >= 2 else low_label
            print(f"  {dim_name}: mean={m:.2f} → 偏向{direction}")

# === Part 4: 综合建议 ===
print("\n" + "=" * 80)
print("五、MBTI整合建议")
print("=" * 80)
print("""
1. 现有Q43已覆盖MBTI的3/4维度(E/I, T/F, J/P), 缺少S/N(感知方式)
2. D/Z(直球/佛系) ≈ T/F(决策方式) 是预测冲突行为的关键维度
3. E/I(社交能量) 预测了对宿舍社交环境的需求强度
4. T/M(秩序) + L/O(作息) ≈ J/P(生活方式) 预测了日常习惯匹配度

后续问卷建议增加:
  - 简化MBTI 4题 (每题二选一, 8题总计)
  - S/N维度: "室友随意动你东西, 你更在意: A.物品被弄乱(S) B.边界被侵犯(N)"
  - 或者直接使用10题TIPI大五人格量表 (覆盖更全面, 学术认可度更高)
""")
