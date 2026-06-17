# -*- coding: utf-8 -*-
"""MBTI交叉分析: 维度×冲突行为×宿舍适应 + 聚类画像可视化数据"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from scipy import stats as sp_stats
from collections import Counter

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

# New MBTI columns: 88=MBTI_type, 89=E/I, 90=S/N, 91=T/F, 92=J/P, 93=personality_cluster
MBTI_TYPE = 88
MBTI_EI = 89    # 1=E, 2=I
MBTI_SN = 90    # 1=S, 2=N
MBTI_TF = 91    # 1=T, 2=F
MBTI_JP = 92    # 1=J, 2=P
CLUSTER = 93    # 聚类标签

resolve = [ws.cell(row, 30).value for row in range(2, N+2)]
anxiety = [ws.cell(row, 23).value for row in range(2, N+2)]
disturbed = [ws.cell(row, 28).value for row in range(2, N+2)]
relation = [ws.cell(row, 31).value for row in range(2, N+2)]
conflict = [ws.cell(row, 25).value for row in range(2, N+2)]
switch = [ws.cell(row, 20).value for row in range(2, N+2)]
schedule = [ws.cell(row, 33).value for row in range(2, N+2)]

mbti_valid = [i for i in range(N) if ws.cell(i+2, MBTI_TYPE).value is not None]
NC = len(mbti_valid)

def mbti_val(col, row_idx):
    return ws.cell(row_idx+2, col).value

print("=" * 80)
print(f"MBTI交叉分析 (N={NC})")
print("=" * 80)

# === 1. MBTI四维度 × 冲突解决方式 ===
print("\n一、MBTI维度 × 冲突解决方式")
print("-" * 50)

for dim_col, dim_name, label0, label1 in [
    (MBTI_EI, 'E/I', 'E外向', 'I内向'),
    (MBTI_SN, 'S/N', 'S实感', 'N直觉'),
    (MBTI_TF, 'T/F', 'T思维', 'F情感'),
    (MBTI_JP, 'J/P', 'J判断', 'P感知'),
]:
    print(f"\n▶ {dim_name} {label0} vs {label1}")
    for gval, gname in [(1, label0), (2, label1)]:
        idxs = [i for i in mbti_valid if mbti_val(dim_col, i) == gval]
        res = [resolve[i] for i in idxs if isinstance(resolve[i], (int,float))]
        cnt = Counter(res)
        total = len(res)
        if total < 3: continue
        direct = cnt.get(1, 0); third = cnt.get(2, 0); endure = cnt.get(4, 0)
        print(f"  {gname} (N={total}): 直接沟通{direct}({direct/total*100:.0f}%) "
              f"第三方{third}({third/total*100:.0f}%) 忍耐{endure}({endure/total*100:.0f}%)")

    # Chi-square 沟通vs忍耐
    g0_r = [resolve[i] for i in mbti_valid if mbti_val(dim_col,i)==1 and isinstance(resolve[i],(int,float))]
    g1_r = [resolve[i] for i in mbti_valid if mbti_val(dim_col,i)==2 and isinstance(resolve[i],(int,float))]
    if len(g0_r) >= 3 and len(g1_r) >= 3:
        g0_comm = sum(1 for r in g0_r if r in [1,2]); g0_end = sum(1 for r in g0_r if r==4)
        g1_comm = sum(1 for r in g1_r if r in [1,2]); g1_end = sum(1 for r in g1_r if r==4)
        if min(g0_comm,g0_end,g1_comm,g1_end) >= 2:
            chi2, p = sp_stats.chi2_contingency([[g0_comm,g0_end],[g1_comm,g1_end]])[:2]
            print(f"  χ2(沟通vs忍耐)={chi2:.2f}, p={p:.4f}{'**' if p<.05 else ' †' if p<.1 else ''}")

# === 2. MBTI四维度 × 宿舍适应指标 ===
print("\n\n二、MBTI维度 × 宿舍适应指标 (Cohen's d)")
print("-" * 50)

for dim_col, dim_name, label0, label1 in [
    (MBTI_EI, 'E/I', 'E外向', 'I内向'),
    (MBTI_TF, 'T/F', 'T思维', 'F情感'),
    (MBTI_JP, 'J/P', 'J判断', 'P感知'),
    (MBTI_SN, 'S/N', 'S实感', 'N直觉'),
]:
    print(f"\n▶ {dim_name}")
    g0 = [i for i in mbti_valid if mbti_val(dim_col,i)==1]
    g1 = [i for i in mbti_valid if mbti_val(dim_col,i)==2]
    for name, col in [('焦虑',23),('被吵醒',28),('关系',31),('冲突',25),('调宿念头',20)]:
        v0 = [ws.cell(i+2,col).value for i in g0 if isinstance(ws.cell(i+2,col).value,(int,float))]
        v1 = [ws.cell(i+2,col).value for i in g1 if isinstance(ws.cell(i+2,col).value,(int,float))]
        if len(v0) >= 3 and len(v1) >= 3:
            d = (np.mean(v0)-np.mean(v1))/np.sqrt((np.std(v0,ddof=1)**2+np.std(v1,ddof=1)**2)/2)
            t, p = sp_stats.ttest_ind(v0, v1)
            sig = '**' if p < .05 else ('†' if p < .1 else '')
            if abs(d) >= 0.2:
                print(f"  {name}: {label0}{np.mean(v0):.2f} vs {label1}{np.mean(v1):.2f}, d={d:+.3f}, p={p:.4f}{sig}")

# === 3. 聚类四型 × 综合对比 ===
print("\n\n三、人格-行为聚类四型 综合画像")
print("=" * 80)

cluster_names = ['高焦虑忍耐型', '理性沟通型', '高痛苦秩序型', '佛系低痛苦型']
for ci, cname in enumerate(cluster_names):
    idxs = [i for i in mbti_valid if mbti_val(CLUSTER, i) == cname]
    if len(idxs) < 2: continue
    print(f"\n▶ Cluster {ci+1}: {cname} (N={len(idxs)})")

    # MBTI分布
    mbtis = [mbti_val(MBTI_TYPE, i) for i in idxs]
    print(f"  MBTI: {dict(Counter(mbtis).most_common(4))}")

    # 核心指标
    for name, col, unit in [('焦虑',23,'天'),('被吵醒',28,'次'),('关系',31,'/5'),
                             ('冲突',25,'次'),('调宿念头',20,'/4'),('作息一致性',33,'/5')]:
        vals = [ws.cell(i+2,col).value for i in idxs if isinstance(ws.cell(i+2,col).value,(int,float))]
        if vals: print(f"  {name}: {np.mean(vals):.2f}{unit}")

    # 冲突方式
    res = [resolve[i] for i in idxs if isinstance(resolve[i],(int,float))]
    if res:
        cnt = Counter(res)
        t = len(res)
        print(f"  冲突解决: 直接{cnt.get(1,0)}({cnt.get(1,0)/t*100:.0f}%) "
              f"第三方{cnt.get(2,0)}({cnt.get(2,0)/t*100:.0f}%) "
              f"忍耐{cnt.get(4,0)}({cnt.get(4,0)/t*100:.0f}%)")

    # 分配建议
    if ci == 0: print(f"  分配策略: 避免与夜猫O型同寝, 必须匹配作息一致且愿意沟通的舍友")
    elif ci == 1: print(f"  分配策略: 可适配多类型, 作为'调解者'角色, 优先匹配高焦虑型提供支持")
    elif ci == 2: print(f"  分配策略: 同型匹配风险高(互相挑剔), 建议配佛系低痛苦型作为缓冲")
    elif ci == 3: print(f"  分配策略: 万能适配型, 弹性大包容性强, 可匹配任何类型")

# === 4. T/F × F型(情感)内部细分 ===
print("\n\n四、F情感型内部细分: 忍耐者的两种面孔")
print("-" * 50)
f_types = [i for i in mbti_valid if mbti_val(MBTI_TF, i) == 2]
print(f"F型(N={len(f_types)}): 偏好和谐, 回避直接冲突")

# F+忍耐 vs F+沟通
f_endure = [i for i in f_types if isinstance(resolve[i],(int,float)) and resolve[i]==4]
f_comm = [i for i in f_types if isinstance(resolve[i],(int,float)) and resolve[i] in [1,2]]
print(f"  F+忍耐(N={len(f_endure)}): 焦虑{np.mean([anxiety[i] for i in f_endure if isinstance(anxiety[i],(int,float))]):.1f}天")
print(f"  F+沟通(N={len(f_comm)}): 焦虑{np.mean([anxiety[i] for i in f_comm if isinstance(anxiety[i],(int,float))]):.1f}天")

# T型对比
t_types = [i for i in mbti_valid if mbti_val(MBTI_TF, i) == 1]
t_endure = [i for i in t_types if isinstance(resolve[i],(int,float)) and resolve[i]==4]
t_comm = [i for i in t_types if isinstance(resolve[i],(int,float)) and resolve[i] in [1,2]]
print(f"  T+忍耐(N={len(t_endure)}): 焦虑{np.mean([anxiety[i] for i in t_endure if isinstance(anxiety[i],(int,float))]):.1f}天")
print(f"  T+沟通(N={len(t_comm)}): 焦虑{np.mean([anxiety[i] for i in t_comm if isinstance(anxiety[i],(int,float))]):.1f}天")

# === 5. L/O作息维度的特殊地位 ===
print("\n\n五、L/O作息节律: 被低估的最强预测维度")
print("-" * 50)
# From Q43: L/O is second letter, col 76 type code
# Decode Q43 L/O from type_map
q43_type_vals = [ws.cell(i+2, 76).value for i in range(N)]
type_map_l = {1,2,3,4,9,10,11,12}  # types with L (晨鸟)
for gname, gset in [('L晨鸟型', type_map_l), ('O夜猫型', set(range(1,17))-type_map_l)]:
    idxs = [i for i in mbti_valid if int(q43_type_vals[i]) in gset]
    if len(idxs) < 3: continue
    anx = [anxiety[i] for i in idxs if isinstance(anxiety[i],(int,float))]
    dis = [disturbed[i] for i in idxs if isinstance(disturbed[i],(int,float))]
    rel = [relation[i] for i in idxs if isinstance(relation[i],(int,float))]
    res = [resolve[i] for i in idxs if isinstance(resolve[i],(int,float))]
    print(f"  {gname} (N={len(idxs)}): 焦虑{np.mean(anx):.1f}天 被吵醒{np.mean(dis):.1f}次 关系{np.mean(rel):.2f}")
    if res:
        cnt = Counter(res)
        print(f"    冲突方式: 直接{cnt.get(1,0)}({cnt.get(1,0)/len(res)*100:.0f}%) "
              f"第三方{cnt.get(2,0)}({cnt.get(2,0)/len(res)*100:.0f}%) "
              f"忍耐{cnt.get(4,0)}({cnt.get(4,0)/len(res)*100:.0f}%)")

# === 6. Dashboard-ready summary ===
print("\n\n" + "=" * 80)
print("六、看板数据摘要 (Dashboard-ready)")
print("=" * 80)

# MBTI type distribution for chart
mbti_dist = Counter([mbti_val(MBTI_TYPE, i) for i in mbti_valid])
top_mbti = mbti_dist.most_common(8)
print("\nMBTI分布(Top8):")
labels = []; data = []
for t, cnt in top_mbti:
    labels.append(t); data.append(cnt)
    print(f"  {t}: {cnt}")
print(f"  Chart labels: {labels}")
print(f"  Chart data: {data}")

# MBTI dimension bar chart data
print(f"\nMBTI四维度分布:")
for dim_col, dim_name, l1, l2 in [(MBTI_EI,'E/I','E外向','I内向'),(MBTI_SN,'S/N','S实感','N直觉'),
                                    (MBTI_TF,'T/F','T思维','F情感'),(MBTI_JP,'J/P','J判断','P感知')]:
    c1 = sum(1 for i in mbti_valid if mbti_val(dim_col,i)==1)
    c2 = sum(1 for i in mbti_valid if mbti_val(dim_col,i)==2)
    print(f"  {dim_name}: {l1}={c1}, {l2}={c2}")

# Cluster distribution
cl_dist = Counter([mbti_val(CLUSTER, i) for i in mbti_valid])
print(f"\n聚类分布:")
for cl, cnt in cl_dist.most_common():
    print(f"  {cl}: {cnt}")

# Key cross-tab: MBTI × Conflict Resolution
print(f"\nT/F × 冲突解决 (关键交叉):")
for tf_val, tf_name in [(1,'T思维'),(2,'F情感')]:
    idxs = [i for i in mbti_valid if mbti_val(MBTI_TF,i)==tf_val]
    res = [resolve[i] for i in idxs if isinstance(resolve[i],(int,float))]
    cnt = Counter(res); t = len(res)
    print(f"  {tf_name}(N={t}): 直接{cnt.get(1,0)}({cnt.get(1,0)/t*100:.0f}%) "
          f"第三方{cnt.get(2,0)}({cnt.get(2,0)/t*100:.0f}%) 忍耐{cnt.get(4,0)}({cnt.get(4,0)/t*100:.0f}%)")

# E/I × 焦虑/被吵醒
print(f"\nE/I × 被吵醒次数:")
for ei_val, ei_name in [(1,'E外向'),(2,'I内向')]:
    idxs = [i for i in mbti_valid if mbti_val(MBTI_EI,i)==ei_val]
    dis = [disturbed[i] for i in idxs if isinstance(disturbed[i],(int,float))]
    anx = [anxiety[i] for i in idxs if isinstance(anxiety[i],(int,float))]
    print(f"  {ei_name}(N={len(idxs)}): 被吵醒{np.mean(dis):.2f}次 焦虑{np.mean(anx):.2f}天")
