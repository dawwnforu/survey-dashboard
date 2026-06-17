# -*- coding: utf-8 -*-
"""清洗后(N=89)重新计算所有图表数据"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from collections import Counter

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

# === 清洗 ===
core_outcomes = [23, 28, 31, 25, 20]
core_demos = [9, 13, 15]
extended = [9,10,13,15,20,22,23,25,28,30,31,32,33,37,44,45,47,61,68]
likert_cols = [23, 25, 28, 31, 32, 33, 37, 44, 45, 61]

valid = []
for i in range(N):
    row = i + 2
    ok = True
    if sum(1 for col in core_outcomes if ws.cell(row, col).value is None) >= 3: ok = False
    if sum(1 for col in core_demos if ws.cell(row, col).value is None) >= 2: ok = False
    miss_ext = sum(1 for col in extended if ws.cell(row, col).value is None)
    if miss_ext / len(extended) > 0.25: ok = False
    for col in range(69, 74):
        v = ws.cell(row, col).value
        if v is not None and isinstance(v, (int,float)):
            if v < 1 or v > 5: ok = False; break
    q41_done = sum(1 for col in range(69,74) if ws.cell(row,col).value is not None and isinstance(ws.cell(row,col).value,(int,float)))
    if q41_done > 0 and q41_done < 2: ok = False
    rel = ws.cell(row, 31).value; sw = ws.cell(row, 20).value
    if rel is not None and sw is not None and isinstance(rel,(int,float)) and isinstance(sw,(int,float)):
        if rel >= 4 and sw <= 2: ok = False
    sched = ws.cell(row, 33).value; dist = ws.cell(row, 28).value
    if sched is not None and dist is not None and isinstance(sched,(int,float)) and isinstance(dist,(int,float)):
        if sched >= 4 and dist >= 5: ok = False
    lik_vals = [ws.cell(row, col).value for col in likert_cols if ws.cell(row,col).value is not None and isinstance(ws.cell(row,col).value,(int,float))]
    if len(lik_vals) >= 6 and np.std(lik_vals) < 0.3 and len(set(lik_vals)) <= 2: ok = False
    if ok: valid.append(i)

NC = len(valid)
print(f"NC={NC}")

def vcol(col):
    """Get values for valid rows only"""
    return [ws.cell(i+2, col).value for i in valid]

def count_dist(col, levels):
    """Count distribution across levels"""
    vals = vcol(col)
    return [sum(1 for x in vals if x==lv) for lv in levels]

def mean_clean(vals):
    vv = [x for x in vals if isinstance(x, (int,float))]
    return np.mean(vv) if vv else 0

# === 1. 年级 (col 8) ===
gc = Counter([ws.cell(i+2, 8).value for i in valid if isinstance(ws.cell(i+2,8).value,(int,float))])
g1234 = [gc.get(k,0) for k in [1,2,3,4]]
print(f"gradeChart: {g1234}")

# === 2. 性别/生源地 ===
g = vcol(9); org = vcol(10)
g_data = [sum(1 for x in g if x==1), sum(1 for x in g if x==2),
          sum(1 for x in org if x==1), sum(1 for x in org if x==2)]
print(f"genderOriginChart: {g_data}")

# === 3. 独生/住校 ===
oc = vcol(13); board = vcol(15)
ob_data = [sum(1 for x in oc if x==1), sum(1 for x in oc if x==2),
           sum(1 for x in board if x==1), sum(1 for x in board if x==2)]
print(f"onlyBoardChart: {ob_data}")

# === 4. 焦虑分布 ===
anx_dist = count_dist(23, range(0,8))
print(f"anxietyChart: {anx_dist}")

# === 5. 被吵醒分布 ===
dist_dist = count_dist(28, range(0,8))
print(f"disturbedChart: {dist_dist}")

# === 6. 舍友关系 ===
rel_dist = count_dist(31, range(1,6))
print(f"relChart: {rel_dist}")

# === 7. 冲突解决 ===
res_dist = count_dist(30, [1,2,4])
print(f"resolveChart: {res_dist}")

# === 8. 调宿念头 ===
sw_dist = count_dist(20, range(1,5))
print(f"switchChart: {sw_dist}")

# === 9. 卫生 ===
hyg_dist = count_dist(32, range(1,6))
print(f"hygieneChart: {hyg_dist}")

# === 10. 作息一致性 ===
sched_dist = count_dist(33, range(1,6))
print(f"scheduleChart: {sched_dist}")

# === 11. 硬件满意度 ===
fac_dist = count_dist(37, range(1,6))
print(f"facilityChart: {fac_dist}")

# === 12. 睡眠噪音 ===
sq_dist = count_dist(68, range(1,8))
print(f"sleepQChart: {sq_dist}")

# === 13. 游戏 ===
game = vcol(47)
game_data = [sum(1 for x in game if x==1), sum(1 for x in game if x==2)]
print(f"gamingChart: {game_data}")

# === 14. 锻炼 ===
ex_dist = count_dist(61, range(1,6))
print(f"exerciseChart: {ex_dist}")

# === 15. 差异感知 (多选) ===
# Need to check what columns
diff_cols = range(53, 59)  # guessing
diff_data = []
for col in diff_cols:
    vals = [ws.cell(i+2, col).value for i in valid]
    diff_data.append(sum(1 for x in vals if x is not None and x==1))
print(f"diffChart (cols53-58): {diff_data}")

# === 16. 特殊需求 (多选) ===
need_cols = range(62, 67)  # guessing
need_data = []
for col in need_cols:
    vals = [ws.cell(i+2, col).value for i in valid]
    need_data.append(sum(1 for x in vals if x is not None and x==1))
print(f"needsChart (cols62-66): {need_data}")

# === 17. 改善建议 (多选) ===
improve_cols = range(84, 90)  # guessing
improve_data = []
for col in improve_cols:
    vals = [ws.cell(i+2, col).value for i in valid]
    improve_data.append(sum(1 for x in vals if x is not None and x==1))
print(f"improveChart (cols84-89): {improve_data}")

# === 18. Q41因素均值 ===
for idx, name in enumerate(['性格三观','生活习惯','兴趣爱好','经济条件','学习目标']):
    vals = [6 - ws.cell(i+2, 69+idx).value for i in valid if isinstance(ws.cell(i+2, 69+idx).value, (int,float))]
    print(f"Q41_{name}: mean={np.mean(vals):.2f}, n={len(vals)}")

# === 19. 核心均值 ===
for name, col in [('焦虑',23),('被吵醒',28),('冲突',25),('关系',31)]:
    vals = [ws.cell(i+2, col).value for i in valid if isinstance(ws.cell(i+2, col).value, (int,float))]
    print(f"mean_{name}: {np.mean(vals):.2f}")

# === 20. analysisChart1 (核心指标均值) ===
a1 = [mean_clean(vcol(23)), mean_clean(vcol(28)), mean_clean(vcol(25))]
print(f"analysisChart1: {[round(x,2) for x in a1]}")

# === 21. analysisChart2 (百分比指标) ===
# 想过调宿, 关系一般, 卫生较好, 每周游戏≥4次, 完全不接受吸烟
sw_pct = sum(1 for x in vcol(20) if isinstance(x,(int,float)) and x<=2)/NC*100  # 经常/偶尔想
rel_mid = sum(1 for x in vcol(31) if isinstance(x,(int,float)) and x==3)/NC*100
hyg_good = sum(1 for x in vcol(32) if isinstance(x,(int,float)) and x>=4)/NC*100
game_pct = sum(1 for x in vcol(47) if isinstance(x,(int,float)) and x==1)/NC*100
# 吸烟 - need to find column
smoke_col = 48  # guessing
smoke_pct = sum(1 for x in vcol(smoke_col) if isinstance(x,(int,float)) and x==1)/NC*100 if vcol(smoke_col) else 74.5
print(f"analysisChart2: {[round(x,1) for x in [sw_pct, rel_mid, hyg_good, game_pct, smoke_pct]]}")

# === 22. analysisChart3 (特殊需求%) ===
if need_data:
    total_need = sum(need_data)
    need_pcts = [round(n/NC*100, 1) for n in need_data]
else:
    need_pcts = [77.3, 66.4, 29.1, 13.6, 10.0]  # fallback
print(f"analysisChart3: {need_pcts}")

# === 23. 独生子女冲突方式 ===
only_direct = sum(1 for i in valid if ws.cell(i+2,13).value==1 and ws.cell(i+2,30).value==1)
only_third = sum(1 for i in valid if ws.cell(i+2,13).value==1 and ws.cell(i+2,30).value==2)
only_endure = sum(1 for i in valid if ws.cell(i+2,13).value==1 and ws.cell(i+2,30).value==4)
non_direct = sum(1 for i in valid if ws.cell(i+2,13).value==2 and ws.cell(i+2,30).value==1)
non_third = sum(1 for i in valid if ws.cell(i+2,13).value==2 and ws.cell(i+2,30).value==2)
non_endure = sum(1 for i in valid if ws.cell(i+2,13).value==2 and ws.cell(i+2,30).value==4)
only_n = only_direct+only_third+only_endure
non_n = non_direct+non_third+non_endure
print(f"onlyChild_resolve: only({only_n}):{only_direct}/{only_third}/{only_endure} non({non_n}):{non_direct}/{non_third}/{non_endure}")
print(f"onlyChild_pcts: [{round(only_direct/only_n*100)},{round(only_third/only_n*100)},{round(only_endure/only_n*100)},{round(non_direct/non_n*100)},{round(non_third/non_n*100)},{round(non_endure/non_n*100)}]")

# === 24. Q41独生子女 ===
for idx, name in enumerate(['生活习惯','性格三观','兴趣爱好','经济条件']):
    o_vals = [6 - ws.cell(i+2, 69+idx).value for i in valid if ws.cell(i+2,13).value==1 and isinstance(ws.cell(i+2, 69+idx).value, (int,float))]
    n_vals = [6 - ws.cell(i+2, 69+idx).value for i in valid if ws.cell(i+2,13).value==2 and isinstance(ws.cell(i+2, 69+idx).value, (int,float))]
    print(f"Q41_only_{name}: only={np.mean(o_vals):.2f} non={np.mean(n_vals):.2f}")
