# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from scipy import stats

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1
def c(col): return [ws.cell(row, col).value for row in range(2, ws.max_row+1)]

boarding = c(15)    # 1=有, 2=无
months   = c(22)    # Q12 累计住校月数
anxiety  = c(23)
disturbed = c(28)
relation = c(31)
conflict = c(25)
gender   = c(9)

print("=" * 70)
print("住校经历(1=有,2=无) × 各指标")
print("=" * 70)
has_exp = [i for i in range(N) if boarding[i]==1]
no_exp  = [i for i in range(N) if boarding[i]==2]
print("有经历N={}, 无经历N={}".format(len(has_exp), len(no_exp)))
for name, vals in [('焦虑',anxiety),('被吵醒',disturbed),('舍友关系',relation),('冲突次数',conflict)]:
    h = [vals[i] for i in has_exp if isinstance(vals[i],(int,float))]
    n = [vals[i] for i in no_exp if isinstance(vals[i],(int,float))]
    t,p = stats.ttest_ind(h,n)
    d = (np.mean(h)-np.mean(n))/np.sqrt((np.std(h,ddof=1)**2+np.std(n,ddof=1)**2)/2)
    print("  {}: 有经历{:.2f} vs 无经历{:.2f}, d={:.3f}, p={:.4f}".format(name, np.mean(h), np.mean(n), d, p))

print()
print("=" * 70)
print("累计住校月数 × 各指标 (按经验分段)")
print("=" * 70)
valid_m = [(i, months[i]) for i in range(N) if isinstance(months[i],(int,float))]
low_m  = [(i,m) for i,m in valid_m if m <= 12]
mid_m  = [(i,m) for i,m in valid_m if 12 < m <= 36]
high_m = [(i,m) for i,m in valid_m if m > 36]
for label, group in [("0-12月(新手)",low_m),("13-36月(中等)",mid_m),("37+月(老手)",high_m)]:
    if len(group) < 3: continue
    idxs = [x[0] for x in group]
    print("  {} N={}: 焦虑{:.1f} 被吵醒{:.1f} 关系{:.1f} 冲突{:.2f}".format(
        label, len(group),
        np.mean([anxiety[i] for i in idxs if isinstance(anxiety[i],(int,float))]),
        np.mean([disturbed[i] for i in idxs if isinstance(disturbed[i],(int,float))]),
        np.mean([relation[i] for i in idxs if isinstance(relation[i],(int,float))]),
        np.mean([conflict[i] for i in idxs if isinstance(conflict[i],(int,float))])
    ))

# Spearman: months vs wellbeing
print()
print("=" * 70)
print("Spearman: 累计住校月数(连续) vs 各指标")
print("=" * 70)
month_vals = [months[i] for i in range(N) if isinstance(months[i],(int,float))]
for name, vals in [('焦虑',anxiety),('被吵醒',disturbed),('舍友关系',relation),('冲突次数',conflict)]:
    pairs = [(months[i], vals[i]) for i in range(N) if isinstance(months[i],(int,float)) and isinstance(vals[i],(int,float))]
    m = [x[0] for x in pairs]
    v = [x[1] for x in pairs]
    rho, p = stats.spearmanr(m, v)
    sig = '**' if p<0.05 else ('†' if p<0.10 else '')
    print("  {}: rho={:.3f}, p={:.4f} {}".format(name, rho, p, sig))

print()
print("=" * 70)
print("性别 × 高/低痛苦组 (确认方向)")
print("=" * 70)
# Recreate pain index roughly
from collections import Counter
print("性别编码: 1=男, 2=女")
print("高痛苦组性别分布:")
# Just compute simple pain: anxiety + disturbed (higher = worse)
pain_scores = []
for i in range(N):
    if isinstance(anxiety[i],(int,float)) and isinstance(disturbed[i],(int,float)) and isinstance(relation[i],(int,float)):
        rel_pain = 6 - relation[i]  # reverse
        pain_scores.append((i, anxiety[i] + disturbed[i] + rel_pain))
pain_scores.sort(key=lambda x: x[1], reverse=True)
n_group = len(pain_scores) // 3
high_pain = pain_scores[:n_group]
low_pain = pain_scores[-n_group:]
high_genders = [gender[i] for i,_ in high_pain if isinstance(gender[i],(int,float))]
low_genders = [gender[i] for i,_ in low_pain if isinstance(gender[i],(int,float))]
print("  高痛苦组: 男={}, 女={}, 均值={:.2f}".format(
    sum(1 for g in high_genders if g==1), sum(1 for g in high_genders if g==2), np.mean(high_genders)))
print("  低痛苦组: 男={}, 女={}, 均值={:.2f}".format(
    sum(1 for g in low_genders if g==1), sum(1 for g in low_genders if g==2), np.mean(low_genders)))
t,p = stats.ttest_ind(high_genders, low_genders)
print("  t={:.2f}, p={:.4f}".format(t,p))
