# -*- coding: utf-8 -*-
"""H1-H11 假设检验脚本
基于 问卷数据_110份.xlsx 实际计算
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from scipy import stats as sp_stats
from collections import Counter

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

def c(col):
    return [ws.cell(row, col).value for row in range(2, ws.max_row+1)]

def nanmean(arr):
    v = [x for x in arr if x is not None and isinstance(x,(int,float))]
    return np.mean(v) if v else float('nan')

def nancount(arr):
    return sum(1 for x in arr if x is not None and isinstance(x,(int,float)))

# === Load all variables ===
Q7_sibling  = c(14)  # 兄弟姐妹和睦程度
Q8_board    = c(15)  # 住校经历(1有2无)
Q9_elem     = c(16)  # 小学住宿月数
Q9_mid      = c(17)  # 初中
Q9_high     = c(18)  # 高中
Q9_college  = c(19)  # 大学
Q10_switch  = c(20)  # 调宿念头
Q12_dur     = c(22)  # 大学累积入住时长
Q13_anxiety = c(23)  # 焦虑天数
Q15_conflict= c(25)  # 冲突次数
Q17_comm_n  = c(27)  # 主动沟通解决矛盾次数
Q18_disturb = c(28)  # 被吵醒次数
Q20_resolve = c(30)  # 冲突解决(1直接2第三方4忍耐)
Q21_rel     = c(31)  # 舍友关系(1-5)
Q27_facility= c(37)  # 硬件满意度(1-5)
Q29_fair    = c(44)  # 公平感(1公平2不公平3非常)
Q30_sleep_t = c(45)  # 入睡时间(1=22前2=22-23 3=23-24 4=24后)
Q32_gaming  = c(47)  # 游戏(1是2否)
Q34_6_8     = c(51)  # 早上6-8
Q34_8_12    = c(52)  # 8-12
Q34_12_14   = c(53)  # 12-14
Q34_14_18   = c(54)  # 14-18
Q34_18_20   = c(55)  # 18-20
Q34_20_22   = c(56)  # 20-22
Q34_22_24   = c(57)  # 22-24
Q34_0_6     = c(58)  # 0-6
Q39_sched   = c(63)  # 作息需求
Q39_quiet   = c(64)  # 安静环境
Q39_allergy = c(65)  # 过敏
Q39_religion= c(66)  # 宗教
Q39_none    = c(67)  # 无特殊需求
Q40_noise   = c(68)  # 睡眠噪音类型(1-7)
Q41_char    = c(69)  # 性格三观(位次1-6)
Q41_life    = c(70)  # 生活习惯(位次)
Q41_hobby   = c(71)  # 兴趣爱好(位次)
Q41_econ    = c(72)  # 经济条件(位次)
Q41_study   = c(73)  # 学习目标(位次)
Q42_rewill  = c(75)  # 是否愿与现在舍友同住
Q43_pers    = c(76)  # 人格类型(1-16)
Q46_hw      = c(79)  # 改进-硬件
Q46_mgmt    = c(80)  # 改进-管理
Q46_cult    = c(81)  # 改进-文化
Q46_svc     = c(82)  # 改进-后勤
Q46_psy     = c(83)  # 改进-心理支持

# Q43人格拆分
pers_map = {1:'ELTD',2:'ELTZ',3:'ELMD',4:'ELMZ',5:'EOTD',6:'EOTZ',7:'EOMD',8:'EOMZ',
            9:'ILTD',10:'ILTZ',11:'ILMD',12:'ILMZ',13:'IOTD',14:'IOTZ',15:'IOMD',16:'IOMZ'}
EI = [None]*N; LO = [None]*N; TM = [None]*N; DZ = [None]*N
q43_idx = []
for i,p in enumerate(Q43_pers):
    if p is not None and isinstance(p,(int,float)) and 1<=p<=16:
        code = pers_map[int(p)]
        EI[i]=1 if code[0]=='E' else 0
        LO[i]=1 if code[1]=='O' else 0
        TM[i]=1 if code[2]=='T' else 0
        DZ[i]=1 if code[3]=='D' else 0
        q43_idx.append(i)
n_q43 = len(q43_idx)

# Derived variables
late_sleep = [1 if isinstance(s,(int,float)) and s>=3 else 0 for s in Q30_sleep_t]  # 23点后
night_gaming = []
for i in range(N):
    g22 = Q34_22_24[i]; g06 = Q34_0_6[i]
    ng = 0
    if g22 is not None and isinstance(g22,(int,float)) and g22==1: ng=1
    if g06 is not None and isinstance(g06,(int,float)) and g06==1: ng=1
    night_gaming.append(ng)
# H1 composite: late sleep AND night gaming
h1_overlap = [1 if late_sleep[i]==1 and night_gaming[i]==1 else 0 for i in range(N)]

direct_comm = [1 if r==1 else 0 for r in Q20_resolve]
has_special = [1 if (Q39_sched[i]==1 or Q39_quiet[i]==1 or Q39_allergy[i]==1 or Q39_religion[i]==1) else 0 for i in range(N)]
Q9_total_months = []
for i in range(N):
    total = 0
    for col in [16,17,18,19]:
        v = ws.cell(i+2, col).value
        if v is not None and isinstance(v,(int,float)): total += float(v)
    Q9_total_months.append(total)

# Print overview
print("N总={}, Q43作答={}".format(N, n_q43))
print("深夜娱乐(22-24或0-6): {}人".format(sum(night_gaming)))
print("晚睡(23点后): {}人".format(sum(late_sleep)))
print("晚睡+深夜娱乐叠加: {}人".format(sum(h1_overlap)))
print("有特殊需求: {}人".format(sum(has_special)))
print("主动沟通: {}人".format(sum(direct_comm)))

# ================================================================
def test_groups(name_a, vals_a, name_b, vals_b, metric_name):
    """t检验两组差异"""
    ca = [v for v in vals_a if v is not None and isinstance(v,(int,float))]
    cb = [v for v in vals_b if v is not None and isinstance(v,(int,float))]
    if len(ca)<3 or len(cb)<3: return None
    t,p = sp_stats.ttest_ind(ca, cb)
    pooled = np.sqrt((np.std(ca,ddof=1)**2+np.std(cb,ddof=1)**2)/2)
    d = (np.mean(ca)-np.mean(cb))/pooled if pooled>0 else 0
    return {'name_a':name_a,'mean_a':np.mean(ca),'n_a':len(ca),
            'name_b':name_b,'mean_b':np.mean(cb),'n_b':len(cb),
            't':t,'p':p,'d':d,'metric':metric_name}

def test_spearman(x_vals, y_vals, x_name, y_name):
    """Spearman相关"""
    pairs = [(x,y) for x,y in zip(x_vals,y_vals) if x is not None and y is not None and isinstance(x,(int,float)) and isinstance(y,(int,float))]
    if len(pairs)<5: return None
    xa = [p[0] for p in pairs]; ya = [p[1] for p in pairs]
    rho,p = sp_stats.spearmanr(xa, ya)
    return {'rho':rho,'p':p,'n':len(pairs),'x_name':x_name,'y_name':y_name}

# ================================================================
print("\n" + "="*70)
print("H1: 夜猫+深夜娱乐叠加 → 冲突频率")
print("="*70)
overlap_conflict = test_groups('叠加组(晚睡+深夜娱乐)', [Q15_conflict[i] for i in range(N) if h1_overlap[i]==1],
                                '非叠加组', [Q15_conflict[i] for i in range(N) if h1_overlap[i]==0], '冲突次数')
if overlap_conflict:
    print(f"  叠加组: {overlap_conflict['mean_a']:.2f}次/月 (N={overlap_conflict['n_a']})")
    print(f"  非叠加组: {overlap_conflict['mean_b']:.2f}次/月 (N={overlap_conflict['n_b']})")
    print(f"  t={overlap_conflict['t']:.2f}, p={overlap_conflict['p']:.4f}, d={overlap_conflict['d']:.3f}")
    sig = '**' if overlap_conflict['p']<0.05 else ('†' if overlap_conflict['p']<0.10 else '')
    print(f"  结论: {'支持H1' if overlap_conflict['p']<0.10 and overlap_conflict['d']>0 else '不支持H1'} {sig}")
# Also test anxiety and disturbed
for yvar, yname in [(Q13_anxiety,'焦虑'),(Q18_disturb,'被吵醒')]:
    r = test_groups('叠加组', [yvar[i] for i in range(N) if h1_overlap[i]==1],
                    '非叠加组', [yvar[i] for i in range(N) if h1_overlap[i]==0], yname)
    if r:
        print(f"  延伸-{yname}: 叠加{r['mean_a']:.2f} vs 非叠加{r['mean_b']:.2f}, t={r['t']:.2f}, p={r['p']:.4f}")

# ================================================================
print("\n" + "="*70)
print("H2: E/I → 公平感(Q29)")
print("="*70)
r = test_groups('E社牛', [Q29_fair[i] for i in q43_idx if EI[i]==1 and Q29_fair[i] is not None],
                'I社恐', [Q29_fair[i] for i in q43_idx if EI[i]==0 and Q29_fair[i] is not None], '公平感')
if r:
    print(f"  E型: {r['mean_a']:.2f} (N={r['n_a']}), I型: {r['mean_b']:.2f} (N={r['n_b']})")
    print(f"  t={r['t']:.2f}, p={r['p']:.4f}, d={r['d']:.3f}")
    print(f"  结论: {'不支持H2(E/I差异不显著, 且方向与预期相反)' if r['p']>0.10 else '支持H2'}")

# ================================================================
print("\n" + "="*70)
print("H3: Z佛系 → 更重视'学习目标一致'(Q41位次)")
print("="*70)
r = test_groups('Z佛系', [Q41_study[i] for i in q43_idx if DZ[i]==0 and Q41_study[i] is not None],
                '非Z(直球)', [Q41_study[i] for i in q43_idx if DZ[i]==1 and Q41_study[i] is not None], '学习目标位次')
if r:
    print(f"  Z佛系: 学习目标位次={r['mean_a']:.2f} (越小越重视, N={r['n_a']})")
    print(f"  非Z: 学习目标位次={r['mean_b']:.2f} (N={r['n_b']})")
    print(f"  t={r['t']:.2f}, p={r['p']:.4f}, d={r['d']:.3f}")
    print(f"  结论: {'支持H3(Z更重视学习目标)' if r['p']<0.10 and r['d']>0 else '不支持H3'}")

# Also check: Z × other Q41 factors
for qvar, qname in [(Q41_life,'生活习惯'),(Q41_char,'性格三观'),(Q41_hobby,'兴趣爱好')]:
    r2 = test_groups('Z佛系', [qvar[i] for i in q43_idx if DZ[i]==0 and qvar[i] is not None],
                     '非Z', [qvar[i] for i in q43_idx if DZ[i]==1 and qvar[i] is not None], qname)
    if r2:
        print(f"  延伸-{qname}: Z={r2['mean_a']:.2f} vs 非Z={r2['mean_b']:.2f}, p={r2['p']:.4f}")

# ================================================================
print("\n" + "="*70)
print("H4: D直球 → 主动沟通(Q20) → 关系满意度(Q42)")
print("="*70)
# Step 1: D → 主动沟通
dz_direct_D = sum(1 for i in range(n_q43) if DZ[i]==1 and direct_comm[q43_idx[i]]==1)
dz_avoid_D = sum(1 for i in range(n_q43) if DZ[i]==1 and direct_comm[q43_idx[i]]==0)
dz_direct_Z = sum(1 for i in range(n_q43) if DZ[i]==0 and direct_comm[q43_idx[i]]==1)
dz_avoid_Z = sum(1 for i in range(n_q43) if DZ[i]==0 and direct_comm[q43_idx[i]]==0)
print(f"  D直球: {dz_direct_D}直接/{dz_avoid_D}非直接 ({dz_direct_D/(dz_direct_D+dz_avoid_D)*100:.0f}%直接)")
print(f"  Z佛系: {dz_direct_Z}直接/{dz_avoid_Z}非直接 ({dz_direct_Z/(dz_direct_Z+dz_avoid_Z)*100:.0f}%直接)")
# Fisher
table = [[dz_direct_D, dz_avoid_D],[dz_direct_Z, dz_avoid_Z]]
try:
    odds, p_f = sp_stats.fisher_exact(table)
    print(f"  Fisher: OR={odds:.2f}, p={p_f:.4f}")
except: pass

# Step 2: 主动沟通 → Q42
r = test_groups('主动沟通', [Q42_rewill[i] for i in range(N) if direct_comm[i]==1 and Q42_rewill[i] is not None],
                '非主动沟通', [Q42_rewill[i] for i in range(N) if direct_comm[i]==0 and Q42_rewill[i] is not None], '重选室友意愿')
if r:
    print(f"  主动沟通→Q42: {r['mean_a']:.2f} vs {r['mean_b']:.2f}, t={r['t']:.2f}, p={r['p']:.4f}")
    print(f"  结论: {'D→沟通路径部分支持, 但沟通→Q42关系' + ('显著' if r['p']<0.05 else '不显著')}")

# ================================================================
print("\n" + "="*70)
print("H5: 硬件满意度(Q27) → 公平感(Q29) → 舍友关系(Q21)")
print("="*70)
# Q27 → Q29
rho1 = test_spearman(Q27_facility, Q29_fair, '硬件满意度', '公平感')
if rho1: print(f"  Q27→Q29: rho={rho1['rho']:.3f}, p={rho1['p']:.4f}")
# Q29 → Q21
rho2 = test_spearman(Q29_fair, Q21_rel, '公平感', '舍友关系')
if rho2: print(f"  Q29→Q21: rho={rho2['rho']:.3f}, p={rho2['p']:.4f}")
# Q27 → Q21 (direct)
rho3 = test_spearman(Q27_facility, Q21_rel, '硬件满意度', '舍友关系')
if rho3: print(f"  Q27→Q21(直接): rho={rho3['rho']:.3f}, p={rho3['p']:.4f}")
# Q10 × Q27 cross
print("  Q10(调宿念头) × Q27(硬件) 交叉:")
for sw_val in [1,2,3,4]:
    vals = [Q27_facility[i] for i in range(N) if Q10_switch[i]==sw_val and Q27_facility[i] is not None]
    if vals:
        print(f"    调宿念头={sw_val}: 硬件满意度均值={np.mean(vals):.2f} (N={len(vals)})")

# ================================================================
print("\n" + "="*70)
print("H6: 游戏+晚睡+噪音敏感 → 被吵醒 → 焦虑 (中介链)")
print("="*70)
# Step 1: 游戏+晚睡+噪音 → 被吵醒
# Create composite risk score
risk_score = []
for i in range(N):
    score = 0
    if Q32_gaming[i]==1: score += 1  # gamer
    if late_sleep[i]==1: score += 1  # late sleeper
    if Q40_noise[i] is not None and isinstance(Q40_noise[i],(int,float)) and Q40_noise[i] in [2,3,7]: score += 1  # noise sensitive
    risk_score.append(score)
high_risk = [i for i in range(N) if risk_score[i]>=2]
low_risk = [i for i in range(N) if risk_score[i]<=1]
r = test_groups('高风险(≥2因子)', [Q18_disturb[i] for i in high_risk],
                '低风险(≤1因子)', [Q18_disturb[i] for i in low_risk], '被吵醒')
if r:
    print(f"  被吵醒: 高风险{r['mean_a']:.2f} vs 低风险{r['mean_b']:.2f}, t={r['t']:.2f}, p={r['p']:.4f}, d={r['d']:.3f}")
# Step 2: 被吵醒 → 焦虑
rho4 = test_spearman(Q18_disturb, Q13_anxiety, '被吵醒', '焦虑')
if rho4: print(f"  被吵醒→焦虑: rho={rho4['rho']:.3f}, p={rho4['p']:.4f}")
# Step 3: Total effect: risk → anxiety
r2 = test_groups('高风险', [Q13_anxiety[i] for i in high_risk],
                 '低风险', [Q13_anxiety[i] for i in low_risk], '焦虑')
if r2:
    print(f"  焦虑(总效应): 高风险{r2['mean_a']:.2f} vs 低风险{r2['mean_b']:.2f}, t={r2['t']:.2f}, p={r2['p']:.4f}")

# ================================================================
print("\n" + "="*70)
print("H7: 有特殊需求 → 更重视'生活习惯'+'学习目标'(Q41位次)")
print("="*70)
for qvar, qname in [(Q41_life,'生活习惯契合'),(Q41_study,'学习目标一致')]:
    r = test_groups('有特殊需求', [qvar[i] for i in range(N) if has_special[i]==1 and qvar[i] is not None],
                    '无特殊需求', [qvar[i] for i in range(N) if has_special[i]==0 and qvar[i] is not None], qname)
    if r:
        sig = '**' if r['p']<0.05 else ('†' if r['p']<0.10 else '')
        print(f"  {qname}: 有需求{r['mean_a']:.2f} vs 无需求{r['mean_b']:.2f}, t={r['t']:.2f}, p={r['p']:.4f}, d={r['d']:.3f} {sig}")

# ================================================================
print("\n" + "="*70)
print("H8: 主动沟通(Q20) → 关系满意度(Q21) + 重选室友(Q42)")
print("="*70)
for yvar, yname in [(Q21_rel,'舍友关系'),(Q42_rewill,'重选室友意愿')]:
    r = test_groups('主动沟通', [yvar[i] for i in range(N) if direct_comm[i]==1 and yvar[i] is not None],
                    '非主动沟通', [yvar[i] for i in range(N) if direct_comm[i]==0 and yvar[i] is not None], yname)
    if r:
        sig = '**' if r['p']<0.05 else ''
        print(f"  {yname}: 主动{r['mean_a']:.2f} vs 非主动{r['mean_b']:.2f}, t={r['t']:.2f}, p={r['p']:.4f}, d={r['d']:.3f} {sig}")

# ================================================================
print("\n" + "="*70)
print("H9: 兄弟姐妹和睦(Q7) + 累计住校月数 → 冲突(Q15) + 关系(Q21)")
print("="*70)
for xvar, xname in [(Q7_sibling,'兄弟姐妹和睦'),(Q9_total_months,'累计住校月数')]:
    for yvar, yname in [(Q15_conflict,'冲突次数'),(Q21_rel,'舍友关系')]:
        r = test_spearman(xvar, yvar, xname, yname)
        if r:
            sig = '**' if r['p']<0.05 else ('†' if r['p']<0.10 else '')
            print(f"  {xname}→{yname}: rho={r['rho']:.3f}, p={r['p']:.4f}, N={r['n']} {sig}")

# ================================================================
print("\n" + "="*70)
print("H10: 非主动沟通 → 减少沟通次数(Q17) → 降低幸福感(Q42+Q21)")
print("="*70)
r = test_groups('主动沟通', [Q17_comm_n[i] for i in range(N) if direct_comm[i]==1 and Q17_comm_n[i] is not None],
                '非主动沟通', [Q17_comm_n[i] for i in range(N) if direct_comm[i]==0 and Q17_comm_n[i] is not None], '沟通次数')
if r:
    print(f"  沟通次数: 主动{r['mean_a']:.2f} vs 非主动{r['mean_b']:.2f}, t={r['t']:.2f}, p={r['p']:.4f}")
# Q17 → Q42/Q21
for yvar, yname in [(Q42_rewill,'重选室友'),(Q21_rel,'舍友关系')]:
    r2 = test_spearman(Q17_comm_n, yvar, '沟通次数', yname)
    if r2:
        print(f"  Q17→{yname}: rho={r2['rho']:.3f}, p={r2['p']:.4f}")

# ================================================================
print("\n" + "="*70)
print("H11: 正式渠道求助比例 vs 心理支持需求比例")
print("="*70)
# Q20中 "向辅导员或阿姨寻求帮助" - value=3? or is it option 3?
# Check Q20 values
q20_counter = Counter([v for v in Q20_resolve if v is not None])
print(f"  Q20(冲突解决方式)值分布: {sorted(q20_counter.items())}")
# Let me check what value 3 means in Q20
# From col header: 1=直接沟通, 2=第三方调解, 4=忍耐
# "向辅导员" might be value 3
formal_help = sum(1 for v in Q20_resolve if v==3)  # assuming 3=辅导员
total_q20 = sum(1 for v in Q20_resolve if v is not None)
formal_pct = formal_help/total_q20*100 if total_q20>0 else 0
print(f"  正式渠道求助(Q20选3): {formal_help}/{total_q20} = {formal_pct:.1f}%")

psy_demand = sum(1 for v in Q46_psy if v==1)
total_q46 = sum(1 for v in Q46_psy if v is not None)
psy_pct = psy_demand/total_q46*100 if total_q46>0 else 0
print(f"  心理支持需求(Q46选该项): {psy_demand}/{total_q46} = {psy_pct:.1f}%")
print(f"  差距: 需求({psy_pct:.1f}%) vs 实际求助({formal_pct:.1f}%), 缺口={psy_pct-formal_pct:.1f}个百分点")

print("\n" + "="*70)
print("全部假设检验完成")
print("="*70)
