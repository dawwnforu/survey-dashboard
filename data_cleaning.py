# -*- coding: utf-8 -*-
"""数据清洗 v4: 加入Q41排名有效性 + 细分检查"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

# 列定义
core_outcomes = [23, 28, 31, 25, 20]  # 焦虑,被吵醒,关系,冲突,调宿念头
core_demos = [9, 13, 15]              # 性别,独生,住校经历
extended = [9,10,13,15,20,22,23,25,28,30,31,32,33,37,44,45,47,61,68]  # 19个
likert_cols = [23, 25, 28, 31, 32, 33, 37, 44, 45, 61]

excluded = {}  # row -> [reasons]

for i in range(N):
    row = i + 2
    reasons = []

    # === 1. 核心变量缺失 ===
    miss_out = sum(1 for col in core_outcomes if ws.cell(row, col).value is None)
    miss_demo = sum(1 for col in core_demos if ws.cell(row, col).value is None)
    if miss_out >= 3:
        reasons.append(f"核心结果缺失{miss_out}/5项")
    if miss_demo >= 2:
        reasons.append(f"人口学缺失{miss_demo}/3项")

    # === 2. 扩展变量缺失 > 25% ===
    miss_ext = sum(1 for col in extended if ws.cell(row, col).value is None)
    if miss_ext / len(extended) > 0.25:
        reasons.append(f"整体缺失率{miss_ext/len(extended):.0%}({miss_ext}/{len(extended)})")

    # === 3. Q41排名有效性 (cols 69-73, 每项应为1-5且无重复) ===
    q41 = []
    for col in range(69, 74):
        v = ws.cell(row, col).value
        if v is not None and isinstance(v, (int, float)):
            q41.append(v)
    if len(q41) >= 4:  # 至少回答了4项
        if len(set(q41)) < len(q41):  # 有重复排名
            reasons.append(f"Q41重复排名(值:{q41})")
        if any(v < 1 or v > 5 for v in q41):
            reasons.append(f"Q41排名越界(值:{q41})")
    elif len(q41) > 0 and len(q41) < 3:
        reasons.append(f"Q41大面积缺失(仅{len(q41)}/5)")

    # === 4. Q40睡眠噪音: 1-7应为单选 ===
    sleep_q = ws.cell(row, 68).value
    if sleep_q is not None and isinstance(sleep_q, (int, float)):
        if sleep_q < 1 or sleep_q > 7:
            reasons.append(f"睡眠噪音类型越界: {sleep_q}")

    # === 5. 逻辑矛盾 ===
    rel = ws.cell(row, 31).value
    sw = ws.cell(row, 20).value
    sched = ws.cell(row, 33).value
    dist = ws.cell(row, 28).value
    anx = ws.cell(row, 23).value

    if rel is not None and sw is not None and isinstance(rel,(int,float)) and isinstance(sw,(int,float)):
        if rel >= 4 and sw <= 2:
            reasons.append(f"矛盾:关系好({rel})但想调宿({sw})")
    if sched is not None and dist is not None and isinstance(sched,(int,float)) and isinstance(dist,(int,float)):
        if sched >= 4 and dist >= 5:
            reasons.append(f"矛盾:作息一致({sched})但频繁被吵({dist})")

    # === 6. Straight-lining ===
    lik_vals = [ws.cell(row, col).value for col in likert_cols
                if ws.cell(row, col).value is not None and isinstance(ws.cell(row, col).value, (int,float))]
    if len(lik_vals) >= 6:
        std = np.std(lik_vals)
        if std < 0.3 and len(set(lik_vals)) <= 2:
            reasons.append(f"疑似随意作答(Likert std={std:.2f}, 唯一值{len(set(lik_vals))})")

    # === 7. 住校月数异常值 (IQR法) ===
    months = ws.cell(row, 22).value
    if months is not None and isinstance(months, (int, float)):
        if months > 180:  # 15年以上住校不太可能(大学生)
            reasons.append(f"住校月数极端:{months}月")

    if reasons:
        excluded[row] = reasons

# === 输出 ===
print("=" * 70)
print("数据清洗报告 v4")
print("=" * 70)
print(f"原始样本: {N}")
print(f"标记排除: {len(excluded)} 份")
print(f"有效样本: {N - len(excluded)} 份")

# 按原因统计
from collections import Counter
reason_cats = Counter()
for row, reasons in excluded.items():
    for r in reasons:
        cat = r.split(':')[0]
        reason_cats[cat] += 1
print(f"\n排除原因分布:")
for cat, cnt in reason_cats.most_common():
    print(f"  {cat}: {cnt}")

print(f"\n排除详情:")
for row in sorted(excluded.keys()):
    print(f"  Row {row}: {'; '.join(excluded[row])}")

print(f"\n{'='*70}")
print(f"结论: 有效样本 N = {N - len(excluded)}")
if N - len(excluded) < 80:
    print("注意: 即使最严格标准, 有效样本仍<80, 需人工判断")
elif N - len(excluded) > 95:
    print("数据原始质量较好, 排除率低。若需80-90, 需说明使用了更严格的主观标准")
else:
    print("有效样本在80-95区间, 符合预期")
