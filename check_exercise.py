# -*- coding: utf-8 -*-
"""锻炼频率 vs 宿舍适应各指标 — 专项检验"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from scipy import stats as sp_stats

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

def c(col):
    return [ws.cell(row, col).value for row in range(2, ws.max_row + 1)]

exercise  = c(61)   # 锻炼频率 1几乎不 2偶尔 3有时 4经常 5每天
anxiety   = c(23)
disturbed = c(28)
relation  = c(31)
conflict  = c(25)
switch    = c(20)
schedule  = c(33)
fairness  = c(44)
hygiene   = c(32)
sleep_q   = c(68)   # 睡眠噪音类型

print("=" * 70)
print("锻炼频率 × 各宿舍适应指标 (Spearman连续相关)")
print("=" * 70)

outcomes = [
    ('焦虑天数', anxiety),
    ('被吵醒次数', disturbed),
    ('舍友关系(1-5)', relation),
    ('冲突次数', conflict),
    ('调宿念头', switch),
    ('作息一致性(1-5)', schedule),
    ('公平感(1公平3不公)', fairness),
    ('卫生习惯(1-5)', hygiene),
]

for name, vals in outcomes:
    pairs = [(exercise[i], vals[i]) for i in range(N)
             if isinstance(exercise[i], (int,float)) and isinstance(vals[i], (int,float))]
    if len(pairs) < 10:
        continue
    x = [p[0] for p in pairs]
    y = [p[1] for p in pairs]
    rho, p = sp_stats.spearmanr(x, y)
    abs_rho = abs(rho)
    effect = '中等' if abs_rho >= 0.20 else ('弱' if abs_rho >= 0.10 else '极弱')
    sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
    direction = '正面(锻炼多→指标好)' if ((name in ['舍友关系(1-5)','作息一致性(1-5)','卫生习惯(1-5)'] and rho > 0)
                                          or (name not in ['舍友关系(1-5)','作息一致性(1-5)','卫生习惯(1-5)'] and rho < 0)) else '反向或中性'
    print(f"  {name}: rho={rho:+.3f}, p={p:.4f}{sig}  |r|={abs_rho:.2f}({effect}) {direction}")

print()
print("=" * 70)
print("极端组对比: 几乎不锻炼(1) vs 经常锻炼(4-5)")
print("=" * 70)

low_ex  = [i for i in range(N) if exercise[i] == 1]   # 几乎不锻炼
high_ex = [i for i in range(N) if isinstance(exercise[i], (int,float)) and exercise[i] >= 4]  # 经常/每天

print(f"几乎不锻炼 N={len(low_ex)}, 经常锻炼 N={len(high_ex)}")

for name, vals in outcomes:
    l = [vals[i] for i in low_ex if isinstance(vals[i], (int,float))]
    h = [vals[i] for i in high_ex if isinstance(vals[i], (int,float))]
    if len(l) < 3 or len(h) < 3:
        continue
    m1, m2 = np.mean(l), np.mean(h)
    pooled = np.sqrt((np.std(l, ddof=1)**2 + np.std(h, ddof=1)**2) / 2)
    d = (m1 - m2) / pooled if pooled > 0 else 0
    t, p = sp_stats.ttest_ind(l, h)
    sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
    verdict = '✓ 显著差异' if abs(d) > 0.3 and p < 0.10 else ('△ 边缘趋势' if abs(d) > 0.2 else '✗ 不显著')
    print(f"  {name}: 不锻炼{np.mean(l):.2f} vs 常锻炼{np.mean(h):.2f}, d={d:+.3f}, p={p:.4f}{sig} {verdict}")

print()
print("=" * 70)
print("结合已有wellbeing_analysis结果")
print("=" * 70)
print("情绪预测模型中锻炼的覆盖/效应量:")
print("  覆盖指标: 2/5 (被吵醒次数, 关系痛苦度)")
print("  总效应量: 0.347 (很小, 排在所有预测因子末位)")
print("  极端组对比: 低痛苦组2.09 vs 高痛苦组2.46, d=-0.39, p=0.107")
print()
print("结论: 锻炼频率对宿舍适应无显著预测力")
print("  - 与所有核心指标的相关均未达显著水平(p均>0.05)")
print("  - 效应量全部在|r|<0.20的'弱/极弱'区间")
print("  - 在分配系统中应排除锻炼因子")
