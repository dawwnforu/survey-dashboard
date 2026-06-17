# -*- coding: utf-8 -*-
"""检查Q43 cols 76-91 实际数据"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active

# Check Q43 columns (76-91, 16 items)
# But wait - maybe the column numbering is different
# Let's check column headers and first few rows
print("Row 2 (first data row) cols 76-91:")
for col in range(76, 92):
    v = ws.cell(2, col).value
    print(f"  Col {col}: {v}")

print("\nRow 3 cols 76-91:")
for col in range(76, 92):
    v = ws.cell(3, col).value
    print(f"  Col {col}: {v}")

# Count non-None in each Q43 column
print("\n\nNon-None counts per Q43 column:")
for col in range(76, 92):
    cnt = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, col).value is not None)
    print(f"  Col {col}: {cnt} non-null")

# What are the unique values?
print("\nUnique values per Q43 column:")
for col in range(76, 92):
    vals = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, col).value
        if v is not None:
            vals.add(v)
    print(f"  Col {col}: {sorted(vals)}")

# Check if Q43 might be at different columns
# The column 76 might actually be something else
# Let's check around col 76 area
print("\n\nRow 1 (header) around col 76:")
for col in range(74, 93):
    v = ws.cell(1, col).value
    print(f"  Col {col}: {v}")
