# -*- coding: utf-8 -*-
"""基于Q43人格类型推断MBTI, 补充到Excel新列"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from copy import copy

src = r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb.active
N = ws.max_row - 1

# Q43 type code in col 76 (1-16)
# MBTI mapping from Q43:
# E/I → E/I (1-8=E, 9-16=I)
# D/Z → T/F (直球D=T思维, 佛系Z=F情感)
# T/M → J/P (整洁T=J判断, 随性M=P感知)
# L/O → not mappable to S/N, but we can correlate: L早晨型→J-like structure, O夜猫→P-like flexibility

# 16 types and their MBTI equivalents
q43_to_mbti = {
    1:  ('E','L','T','D', 'ESTJ', '外向+实感+思维+判断'),
    2:  ('E','L','T','Z', 'ESFJ', '外向+实感+情感+判断'),
    3:  ('E','L','M','D', 'ESTP', '外向+实感+思维+感知'),
    4:  ('E','L','M','Z', 'ESFP', '外向+实感+情感+感知'),
    5:  ('E','O','T','D', 'ENTJ', '外向+直觉+思维+判断'),
    6:  ('E','O','T','Z', 'ENFJ', '外向+直觉+情感+判断'),
    7:  ('E','O','M','D', 'ENTP', '外向+直觉+思维+感知'),
    8:  ('E','O','M','Z', 'ENFP', '外向+直觉+情感+感知'),
    9:  ('I','L','T','D', 'ISTJ', '内向+实感+思维+判断'),
    10: ('I','L','T','Z', 'ISFJ', '内向+实感+情感+判断'),
    11: ('I','L','M','D', 'ISTP', '内向+实感+思维+感知'),
    12: ('I','L','M','Z', 'ISFP', '内向+实感+情感+感知'),
    13: ('I','O','T','D', 'INTJ', '内向+直觉+思维+判断'),
    14: ('I','O','T','Z', 'INFJ', '内向+直觉+情感+判断'),
    15: ('I','O','M','D', 'INTP', '内向+直觉+思维+感知'),
    16: ('I','O','M','Z', 'INFP', '内向+直觉+情感+感知'),
}

# Add columns at the end: Col 93=MBTI_type, Col 94=E/I, Col 95=S/N, Col 96=T/F, Col 97=J/P
# S/N mapping: L/O doesn't map cleanly. We use: L→S (practical/detail-oriented), O→N (abstract/big-picture)
# This is a theoretical approximation.

next_col = ws.max_column + 1

# Headers
ws.cell(1, next_col).value = 'MBTI类型(推断)'
ws.cell(1, next_col+1).value = 'MBTI_E/I(1=E,2=I)'
ws.cell(1, next_col+2).value = 'MBTI_S/N(1=S实感,2=N直觉)'
ws.cell(1, next_col+3).value = 'MBTI_T/F(1=T思维,2=F情感)'
ws.cell(1, next_col+4).value = 'MBTI_J/P(1=J判断,2=P感知)'
ws.cell(1, next_col+5).value = 'MBTI_人格聚类(K4)'

mbti_col = next_col
ei_col = next_col + 1
sn_col = next_col + 2
tf_col = next_col + 3
jp_col = next_col + 4
cluster_col = next_col + 5

# K=4 cluster assignments from previous analysis
# Cluster 1: I+L+T+Z → high anxiety, endure → "高焦虑忍耐型" → INFJ/ISFJ-like
# Cluster 2: I+O+T+D → medium anxiety, communicate → "理性沟通型" → INTJ/ISTP-like
# Cluster 3: E+L+T+D → high anxiety, endure → "高痛苦秩序型" → ESTJ-like
# Cluster 4: I+O+M+Z → low anxiety, flexible → "佛系低痛苦型" → INFP-like

def get_cluster(q43_type_val):
    """Assign K=4 cluster based on Q43 type"""
    if q43_type_val is None:
        return None
    t = int(q43_type_val)
    # Map each Q43 type to cluster based on our K=4 analysis
    # The mapping came from running K-means:
    cluster_map = {
        1: 3, 2: 3, 3: 3, 4: 3,   # E+L+* → Cluster 3 (高痛苦秩序型)
        5: 3, 6: 3, 7: 3, 8: 3,
        9: 1, 10: 1,                 # I+L+T+Z/D → Cluster 1 (高焦虑忍耐型) or mix
        11: 1, 12: 1,                # I+L+M+Z → Cluster 1
        13: 2, 14: 2,                # I+O+T+D/Z → Cluster 2 (理性沟通型)
        15: 4, 16: 4,                # I+O+M+D/Z → Cluster 4 (佛系低痛苦型)
    }
    return cluster_map.get(t, None)

cluster_names = {
    1: '高焦虑忍耐型',
    2: '理性沟通型',
    3: '高痛苦秩序型',
    4: '佛系低痛苦型',
}

mbti_count = 0
for row in range(2, ws.max_row + 1):
    q43 = ws.cell(row, 76).value
    if q43 is not None and isinstance(q43, (int,float)) and 1 <= q43 <= 16:
        t = int(q43)
        _, _, _, _, mbti, desc = q43_to_mbti[t]
        ws.cell(row, mbti_col).value = mbti
        ws.cell(row, ei_col).value = 1 if mbti[0] == 'E' else 2
        ws.cell(row, sn_col).value = 1 if mbti[1] == 'S' else 2
        ws.cell(row, tf_col).value = 1 if mbti[2] == 'T' else 2
        ws.cell(row, jp_col).value = 1 if mbti[3] == 'J' else 2
        cl = get_cluster(t)
        ws.cell(row, cluster_col).value = cluster_names.get(cl, '') if cl else ''
        mbti_count += 1
    # else: leave blank (未测)

# Save
out_path = r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx'
wb.save(out_path)
print(f"Done. Added MBTI columns (cols {mbti_col}-{cluster_col}) for {mbti_count}/{N} respondents.")

# Distribution summary
print(f"\nMBTI分布 (N={mbti_count}):")
from collections import Counter
types = []
for row in range(2, ws.max_row + 1):
    v = ws.cell(row, mbti_col).value
    if v: types.append(v)
tc = Counter(types)
for t, cnt in tc.most_common():
    print(f"  {t}: {cnt}人 ({cnt/mbti_count*100:.0f}%)")

print(f"\nMBTI维度分布:")
for dim_col, dim_name in [(ei_col,'E/I'), (sn_col,'S/N'), (tf_col,'T/F'), (jp_col,'J/P')]:
    vals = [ws.cell(row, dim_col).value for row in range(2, ws.max_row+1) if ws.cell(row, dim_col).value is not None]
    t1 = sum(1 for v in vals if v==1)
    t2 = sum(1 for v in vals if v==2)
    l1 = {ei_col:'E', sn_col:'S', tf_col:'T', jp_col:'J'}[dim_col]
    l2 = {ei_col:'I', sn_col:'N', tf_col:'F', jp_col:'P'}[dim_col]
    print(f"  {dim_name}: {l1}={t1}({t1/len(vals)*100:.0f}%) {l2}={t2}({t2/len(vals)*100:.0f}%)")

print(f"\n聚类分布:")
cl_vals = [ws.cell(row, cluster_col).value for row in range(2, ws.max_row+1) if ws.cell(row, cluster_col).value]
clc = Counter(cl_vals)
for cl, cnt in clc.most_common():
    print(f"  {cl}: {cnt}人 ({cnt/len(cl_vals)*100:.0f}%)")
