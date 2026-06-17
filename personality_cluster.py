# -*- coding: utf-8 -*-
"""Q43人格类型解码 + MBTI映射 + K-means聚类分析"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np
from scipy import stats as sp_stats
from collections import Counter

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

def c(col):
    return [ws.cell(row, col).value for row in range(2, ws.max_row + 1)]

# Q43 type code in col 76: 1-16
# 16 types from dashboard:
type_map = {
    1:  ('E','L','T','D', 'ELTD','阳光舍长'),
    2:  ('E','L','T','Z', 'ELTZ','温和管家'),
    3:  ('E','L','M','D', 'ELMD','早起冲锋'),
    4:  ('E','L','M','Z', 'ELMZ','佛系早鸟'),
    5:  ('E','O','T','D', 'EOTD','深夜社交'),
    6:  ('E','O','T','Z', 'EOTZ','夜店气氛'),
    7:  ('E','O','M','D', 'EOMD','自由灵魂'),
    8:  ('E','O','M','Z', 'EOMZ','随缘躺平'),
    9:  ('I','L','T','D', 'ILTD','安静领袖'),
    10: ('I','L','T','Z', 'ILTZ','宁静致远'),
    11: ('I','L','M','D', 'ILMD','独行侠'),
    12: ('I','L','M','Z', 'ILMZ','隐士高人'),
    13: ('I','O','T','D', 'IOTD','深夜思考'),
    14: ('I','O','T','Z', 'IOTZ','夜行侠客'),
    15: ('I','O','M','D', 'IOMD','独立艺术'),
    16: ('I','O','M','Z', 'IOMZ','终极躺平'),
}

# Decode Q43 type to dimensions
q43_type = c(76)  # 1-16 or None
E_I_bin = []  # 0=E社牛, 1=I社恐
L_O_bin = []  # 0=L晨鸟, 1=O夜猫
T_M_bin = []  # 0=T整洁, 1=M随性
D_Z_bin = []  # 0=D直球, 1=Z佛系
q43_valid = []  # indices with valid Q43

for i in range(N):
    t = q43_type[i]
    if t is not None and isinstance(t, (int,float)) and 1 <= t <= 16:
        t = int(t)
        info = type_map[t]
        E_I_bin.append(0 if info[0] == 'E' else 1)
        L_O_bin.append(0 if info[1] == 'L' else 1)
        T_M_bin.append(0 if info[2] == 'T' else 1)
        D_Z_bin.append(0 if info[3] == 'D' else 1)
        q43_valid.append(i)
    else:
        E_I_bin.append(None)
        L_O_bin.append(None)
        T_M_bin.append(None)
        D_Z_bin.append(None)

NQ = len(q43_valid)
print(f"Q43有效样本: {NQ}")
print(f"维度分布:")
for name, arr in [('E/I(0=社牛,1=社恐)', E_I_bin), ('L/O(0=晨鸟,1=夜猫)', L_O_bin),
                   ('T/M(0=整洁,1=随性)', T_M_bin), ('D/Z(0=直球,1=佛系)', D_Z_bin)]:
    vals = [v for v in arr if v is not None]
    e_count = sum(1 for v in vals if v == 0)
    i_count = sum(1 for v in vals if v == 1)
    print(f"  {name}: type0={e_count}({e_count/len(vals)*100:.0f}%) type1={i_count}({i_count/len(vals)*100:.0f}%)")

# ============================================================
# Part 1: 人格维度 × 冲突解决方式
# ============================================================
print("\n" + "=" * 70)
print("一、Q43人格维度 × 冲突解决方式")
print("=" * 70)

resolve = c(30)  # 1直接 2第三方 4忍耐
anxiety = c(23)
disturbed = c(28)
relation = c(31)
conflict = c(25)
switch = c(20)
schedule = c(33)

for dim_name, dim_bin, label0, label1 in [
    ('D/Z 冲突模式', D_Z_bin, '直球D型', '佛系Z型'),
    ('E/I 社交能量', E_I_bin, '社牛E型', '社恐I型'),
    ('T/M 空间秩序', T_M_bin, '整洁T型', '随性M型'),
    ('L/O 作息节律', L_O_bin, '晨鸟L型', '夜猫O型'),
]:
    print(f"\n▶ {dim_name} × 冲突解决方式")
    g0 = [i for i in q43_valid if dim_bin[i] == 0]
    g1 = [i for i in q43_valid if dim_bin[i] == 1]

    for gname, gidx in [(label0, g0), (label1, g1)]:
        res = [resolve[i] for i in gidx if isinstance(resolve[i], (int,float))]
        cnt = Counter(res)
        total = len(res)
        if total < 3: continue
        direct = cnt.get(1, 0)
        third = cnt.get(2, 0)
        endure = cnt.get(4, 0)
        print(f"  {gname} (N={total}): 直接沟通{direct}({direct/total*100:.0f}%) "
              f"第三方调解{third}({third/total*100:.0f}%) "
              f"忍耐{endure}({endure/total*100:.0f}%)")

    # Chi-square
    r0 = [resolve[i] for i in g0 if isinstance(resolve[i], (int,float))]
    r1 = [resolve[i] for i in g1 if isinstance(resolve[i], (int,float))]
    if len(r0) >= 3 and len(r1) >= 3:
        l_comm = sum(1 for r in r0 if r in [1,2])
        l_end = sum(1 for r in r0 if r == 4)
        h_comm = sum(1 for r in r1 if r in [1,2])
        h_end = sum(1 for r in r1 if r == 4)
        if min(l_comm, l_end, h_comm, h_end) >= 2:
            table = [[l_comm, l_end], [h_comm, h_end]]
            chi2, p, dof, exp = sp_stats.chi2_contingency(table)
            sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
            print(f"  卡方(沟通vs忍耐): chi2={chi2:.2f}, p={p:.4f}{sig} {label0}→{label1}方向")

# ============================================================
# Part 2: INFP-like vs ESTJ-like
# ============================================================
print("\n" + "=" * 70)
print("二、MBTI模拟: INFP型 vs ESTJ型 对比")
print("=" * 70)

# INFP ≈ I(社恐) + F(佛系D/Z=1, 近似F型和谐优先) + P(随性T/M=1)
# Actually: INFP = I + N + F + P
# Q43 mapping: I型(E_I=1), N型(not measured), F型(D_Z=1佛系≈回避冲突), P型(T_M=1随性)
infp_like = [i for i in q43_valid if E_I_bin[i]==1 and D_Z_bin[i]==1 and T_M_bin[i]==1]
# ESTJ ≈ E + S + T + J
estj_like = [i for i in q43_valid if E_I_bin[i]==0 and D_Z_bin[i]==0 and T_M_bin[i]==0]

print(f"INFP型(I+佛系+随性): N={len(infp_like)}")
print(f"ESTJ型(E+直球+整洁): N={len(estj_like)}")

for gname, gidx, gemoji in [('INFP型', infp_like, '🌿'), ('ESTJ型', estj_like, '📋')]:
    if len(gidx) < 1: continue
    anx = [anxiety[i] for i in gidx if isinstance(anxiety[i], (int,float))]
    dis = [disturbed[i] for i in gidx if isinstance(disturbed[i], (int,float))]
    rel = [relation[i] for i in gidx if isinstance(relation[i], (int,float))]
    con = [conflict[i] for i in gidx if isinstance(conflict[i], (int,float))]
    sw = [switch[i] for i in gidx if isinstance(switch[i], (int,float))]
    res = [resolve[i] for i in gidx if isinstance(resolve[i], (int,float))]
    print(f"  {gemoji} {gname} N={len(gidx)}:")
    print(f"    焦虑: {np.mean(anx):.2f}天/周  被吵醒: {np.mean(dis):.2f}次  关系: {np.mean(rel):.2f}/5")
    print(f"    冲突: {np.mean(con):.2f}次/月  调宿念头: {np.mean(sw):.2f}/4")
    if res:
        cnt = Counter(res)
        print(f"    冲突方式: 直接沟通{cnt.get(1,0)}({cnt.get(1,0)/len(res)*100:.0f}%) "
              f"第三方调解{cnt.get(2,0)}({cnt.get(2,0)/len(res)*100:.0f}%) "
              f"忍耐{cnt.get(4,0)}({cnt.get(4,0)/len(res)*100:.0f}%)")

# t-tests
if len(infp_like) >= 3 and len(estj_like) >= 3:
    print(f"\n  INFP vs ESTJ 统计检验:")
    for name, col in [('焦虑',23),('被吵醒',28),('关系',31),('冲突',25),('调宿念头',20)]:
        iv = [ws.cell(i+2,col).value for i in infp_like if isinstance(ws.cell(i+2,col).value,(int,float))]
        ev = [ws.cell(i+2,col).value for i in estj_like if isinstance(ws.cell(i+2,col).value,(int,float))]
        if len(iv) >= 3 and len(ev) >= 3:
            d = (np.mean(iv)-np.mean(ev))/np.sqrt((np.std(iv,ddof=1)**2+np.std(ev,ddof=1)**2)/2)
            t, p = sp_stats.ttest_ind(iv, ev)
            sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
            direction = 'INFP更高' if d > 0 else 'ESTJ更高'
            print(f"    {name}: d={d:+.3f}, p={p:.4f}{sig} {direction}")

# ============================================================
# Part 3: K-Means聚类分析 (人格 × 行为)
# ============================================================
print("\n" + "=" * 70)
print("三、K-Means聚类: 人格+行为综合画像")
print("=" * 70)

# Feature matrix: 4 personality dimensions + 5 behavioral outcomes
# Only for Q43 valid cases
features = []
feature_names = ['E/I(社恐)','L/O(夜猫)','T/M(随性)','D/Z(佛系)',
                 '焦虑','被吵醒','关系(反向)','冲突','调宿念头']
row_indices = []

for i in q43_valid:
    row = i + 2
    # Personality (binary 0/1)
    pers = [E_I_bin[i], L_O_bin[i], T_M_bin[i], D_Z_bin[i]]
    # Behavioral
    anx = ws.cell(row, 23).value
    dis = ws.cell(row, 28).value
    rel = ws.cell(row, 31).value  # reverse: higher=better
    con = ws.cell(row, 25).value
    sw = ws.cell(row, 20).value

    behav = [anx, dis, rel, con, sw]
    if all(v is not None and isinstance(v, (int,float)) for v in behav):
        # Normalize behavioral to roughly same scale as personality (0-1)
        # anxiety: 0-7 → /7, disturbed: 0-7 → /7, relation: 1-5 → (6-v)/5, conflict: 0-5 → /5, switch: 1-4 → (v-1)/3
        behav_norm = [
            anx/7,
            dis/7,
            (6-rel)/5,  # reverse: higher=worse relationship
            con/5,
            (sw-1)/3
        ]
        features.append(pers + behav_norm)
        row_indices.append(i)

X = np.array(features)
print(f"聚类样本: {len(X)} 人 (Q43完整+核心变量齐全)")
print(f"特征维度: {X.shape[1]} ({', '.join(feature_names)})")

# Standardize
# Manual K-means implementation (no scipy cluster dependency)
def simple_kmeans(X, k, max_iter=100):
    n = X.shape[0]
    # Init centroids randomly from data
    np.random.seed(42)
    idx = np.random.choice(n, k, replace=False)
    centroids = X[idx].copy()

    for _ in range(max_iter):
        # Assign to nearest centroid
        distances = np.zeros((n, k))
        for j in range(k):
            distances[:, j] = np.sum((X - centroids[j])**2, axis=1)
        labels = np.argmin(distances, axis=1)

        # Update centroids
        new_centroids = np.zeros_like(centroids)
        for j in range(k):
            if np.sum(labels == j) > 0:
                new_centroids[j] = X[labels == j].mean(axis=0)
            else:
                new_centroids[j] = centroids[j]

        if np.allclose(centroids, new_centroids):
            break
        centroids = new_centroids

    return labels, centroids

# Try k=3,4,5
for k in [3, 4, 5]:
    labels, centroids = simple_kmeans(X, k)

    print(f"\n{'='*50}")
    print(f"K={k} 聚类结果")
    print(f"{'='*50}")

    for j in range(k):
        mask = labels == j
        n_cluster = np.sum(mask)
        if n_cluster == 0: continue

        print(f"\n  Cluster {j+1} (N={n_cluster})")

        # Personality profile
        pers_mean = X[mask][:, :4].mean(axis=0)
        e_i_label = '社恐I' if pers_mean[0] > 0.5 else '社牛E'
        l_o_label = '夜猫O' if pers_mean[1] > 0.5 else '晨鸟L'
        t_m_label = '随性M' if pers_mean[2] > 0.5 else '整洁T'
        d_z_label = '佛系Z' if pers_mean[3] > 0.5 else '直球D'
        print(f"    人格画像: {e_i_label} + {l_o_label} + {t_m_label} + {d_z_label}")
        print(f"    维度分: E/I={pers_mean[0]:.2f} L/O={pers_mean[1]:.2f} T/M={pers_mean[2]:.2f} D/Z={pers_mean[3]:.2f}")

        # Behavioral profile
        behav_mean = X[mask][:, 4:].mean(axis=0)
        print(f"    焦虑={behav_mean[0]*7:.1f}天  被吵醒={behav_mean[1]*7:.1f}次  "
              f"关系痛苦={behav_mean[2]*5:.1f}  冲突={behav_mean[3]*5:.1f}次  调宿={behav_mean[4]*3+1:.1f}")

        # Conflict resolution style
        cluster_rows = [row_indices[i] for i in range(len(row_indices)) if labels[i] == j]
        res_styles = [resolve[i] for i in cluster_rows if isinstance(resolve[i], (int,float))]
        if res_styles:
            cnt = Counter(res_styles)
            print(f"    冲突方式: 直接沟通{cnt.get(1,0)}({cnt.get(1,0)/len(res_styles)*100:.0f}%) "
                  f"第三方调解{cnt.get(2,0)}({cnt.get(2,0)/len(res_styles)*100:.0f}%) "
                  f"忍耐{cnt.get(4,0)}({cnt.get(4,0)/len(res_styles)*100:.0f}%)")

    # Silhouette-like score (simplified)
    inertia = np.sum(np.min(np.array([np.sum((X - centroids[j])**2, axis=1) for j in range(k)]), axis=0))
    print(f"\n  Inertia (within-cluster SS): {inertia:.2f}")

# ============================================================
# Part 4: 聚类命名建议
# ============================================================
print("\n" + "=" * 70)
print("四、MBTI-宿舍行为 聚类画像总结")
print("=" * 70)
print("""
基于人格维度(E/I, L/O, T/M, D/Z) + 行为指标(焦虑/被吵醒/关系/冲突/调宿)的K-Means聚类:

预期3-4类典型画像:
  Cluster A "高敏感忍耐型": I型 + 佛系Z + 高焦虑 + 偏好忍耐 → MBTI类似INFP/INFJ
  Cluster B "主动沟通型": E型 + 直球D + 低焦虑 + 直接沟通 → MBTI类似ESTJ/ENTJ
  Cluster C "社牛随性型": E型 + 随性M + 中等焦虑 + 第三方调解 → MBTI类似ESFP/ENFP
  Cluster D "秩序焦虑型": T型 + 直球D + 高焦虑 + 频繁冲突 → MBTI类似ISTJ/INTJ

分配系统应用: 同聚类内匹配(相似型)减少摩擦, 或异聚类互补(E/I互补, 直球配佛系)
""")
