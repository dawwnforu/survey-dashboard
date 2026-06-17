# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import numpy as np
from scipy import stats as sp_stats

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

def c(col):
    return [ws.cell(row, col).value for row in range(2, ws.max_row + 1)]

only_child = c(13)   # 1=独生, 2=非独生
origin     = c(10)   # 1=城镇, 2=农村
resolve    = c(30)   # 1=直接, 2=第三方, 4=忍耐
conflict   = c(25)
relation   = c(31)
anxiety    = c(23)
disturbed  = c(28)
switch_idea = c(20)
factor_life  = c(70)
factor_char  = c(69)
factor_hobby = c(71)
factor_econ  = c(72)
factor_study = c(73)

print("=" * 70)
print("独生子女(N=110) × 冲突处理方式、沟通行为、Q41偏好")
print("=" * 70)

# 1. 独生子女 × 冲突解决方式
only_direct = sum(1 for i in range(N) if only_child[i]==1 and resolve[i]==1)
only_third  = sum(1 for i in range(N) if only_child[i]==1 and resolve[i]==2)
only_endure = sum(1 for i in range(N) if only_child[i]==1 and resolve[i]==4)
non_direct  = sum(1 for i in range(N) if only_child[i]==2 and resolve[i]==1)
non_third   = sum(1 for i in range(N) if only_child[i]==2 and resolve[i]==2)
non_endure  = sum(1 for i in range(N) if only_child[i]==2 and resolve[i]==4)
only_n = only_direct+only_third+only_endure
non_n  = non_direct+non_third+non_endure
print("\n▶ 独生子女 × 冲突解决方式")
print("  独生(N={}): 直接{} ({}%)  第三方{} ({}%)  忍耐{} ({}%)".format(
    only_n, only_direct, round(only_direct/only_n*100), only_third, round(only_third/only_n*100), only_endure, round(only_endure/only_n*100)))
print("  非独生(N={}): 直接{} ({}%)  第三方{} ({}%)  忍耐{} ({}%)".format(
    non_n, non_direct, round(non_direct/non_n*100), non_third, round(non_third/non_n*100), non_endure, round(non_endure/non_n*100)))
table = [[only_direct+only_third, only_endure], [non_direct+non_third, non_endure]]
chi2, p, dof, exp = sp_stats.chi2_contingency(table)
print("  卡方(沟通vs忍耐): chi2={:.2f}, p={:.4f}".format(chi2, p))

# 2. 独生子女 × 定量指标
print("\n▶ 独生子女 × 各定量指标 (独立样本t检验)")
for name, vals in [('冲突次数',conflict),('舍友关系(1-5)',relation),('焦虑天数',anxiety),('被吵醒次数',disturbed),('调宿念头',switch_idea)]:
    only_vals = [vals[i] for i in range(N) if only_child[i]==1 and isinstance(vals[i],(int,float))]
    non_vals  = [vals[i] for i in range(N) if only_child[i]==2 and isinstance(vals[i],(int,float))]
    m1, m2 = np.mean(only_vals), np.mean(non_vals)
    pooled = np.sqrt((np.std(only_vals,ddof=1)**2+np.std(non_vals,ddof=1)**2)/2)
    d = (m1-m2)/pooled if pooled>0 else 0
    t, p = sp_stats.ttest_ind(only_vals, non_vals)
    sig = '**' if p<0.05 else ('†' if p<0.10 else '')
    print("  {}: 独生{:.2f} vs 非独生{:.2f}, d={:.3f}, t={:.2f}, p={:.4f} {}".format(name, m1, m2, d, t, p, sig))

# 3. 独生子女 × Q41偏好
print("\n▶ 独生子女 × Q41因素重要度位次(越小越重视)")
for name, vals in [('生活习惯',factor_life),('性格三观',factor_char),('兴趣爱好',factor_hobby),('经济条件',factor_econ),('学习目标',factor_study)]:
    only_vals = [vals[i] for i in range(N) if only_child[i]==1 and isinstance(vals[i],(int,float))]
    non_vals  = [vals[i] for i in range(N) if only_child[i]==2 and isinstance(vals[i],(int,float))]
    m1, m2 = np.mean(only_vals), np.mean(non_vals)
    pooled = np.sqrt((np.std(only_vals,ddof=1)**2+np.std(non_vals,ddof=1)**2)/2)
    d = (m1-m2)/pooled if pooled>0 else 0
    t, p = sp_stats.ttest_ind(only_vals, non_vals)
    sig = '**' if p<0.05 else ('†' if p<0.10 else '')
    print("  {}: 独生{:.2f} vs 非独生{:.2f}, d={:.3f}, t={:.2f}, p={:.4f} {}".format(name, m1, m2, d, t, p, sig))

print("\n" + "=" * 70)
print("生源地(城镇/农村) × 幸福感指标")
print("=" * 70)
urban_n = sum(1 for o in origin if o==1)
rural_n = sum(1 for o in origin if o==2)
print("城镇N={}, 农村N={}".format(urban_n, rural_n))
for name, vals in [('冲突次数',conflict),('舍友关系',relation),('焦虑天数',anxiety),('被吵醒次数',disturbed),('调宿念头',switch_idea)]:
    urban_vals = [vals[i] for i in range(N) if origin[i]==1 and isinstance(vals[i],(int,float))]
    rural_vals = [vals[i] for i in range(N) if origin[i]==2 and isinstance(vals[i],(int,float))]
    m1, m2 = np.mean(urban_vals), np.mean(rural_vals)
    pooled = np.sqrt((np.std(urban_vals,ddof=1)**2+np.std(rural_vals,ddof=1)**2)/2)
    d = (m1-m2)/pooled if pooled>0 else 0
    t, p = sp_stats.ttest_ind(urban_vals, rural_vals)
    sig = '**' if p<0.05 else ('†' if p<0.10 else '')
    print("  {}: 城镇{:.2f} vs 农村{:.2f}, d={:.3f}, t={:.2f}, p={:.4f} {}".format(name, m1, m2, d, t, p, sig))

print("\n" + "=" * 70)
print("关键发现")
print("=" * 70)
print("1. 经济条件(家庭收入/月生活费): 原始问卷中未作为独立变量采集")
print("   Col 72是Q41'经济条件相当'的排序位次(主观重要性),非客观经济水平")
print("   '经济条件与幸福感无关联'是基于Q41排序的延伸推断,未做回归检验")
print("2. 地域文化差异: 问卷中仅采集了'生源地'(城镇/农村),无地域文化变量")
print("3. 独生子女: 以上检验覆盖了冲突方式、沟通行为、Q41偏好全维度")
