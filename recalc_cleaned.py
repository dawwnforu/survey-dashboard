# -*- coding: utf-8 -*-
"""用清洗后数据(N=79)重新计算所有图表数据"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, numpy as np

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active
N = ws.max_row - 1

def c(col):
    return [ws.cell(row, col).value for row in range(2, ws.max_row + 1)]

# === 清洗标准 ===
core_outcomes = [23, 28, 31, 25, 20]
core_demos = [9, 13, 15]
extended = [9,10,13,15,20,22,23,25,28,30,31,32,33,37,44,45,47,61,68]
likert_cols = [23, 25, 28, 31, 32, 33, 37, 44, 45, 61]

valid = []
for i in range(N):
    row = i + 2
    ok = True

    # 核心结果缺失 >=3
    if sum(1 for col in core_outcomes if ws.cell(row, col).value is None) >= 3:
        ok = False
    # 人口学缺失 >=2
    if sum(1 for col in core_demos if ws.cell(row, col).value is None) >= 2:
        ok = False
    # 扩展缺失 >25%
    miss_ext = sum(1 for col in extended if ws.cell(row, col).value is None)
    if miss_ext / len(extended) > 0.25:
        ok = False
    # Q41排名越界 (包含6等无效值)
    for col in range(69, 74):
        v = ws.cell(row, col).value
        if v is not None and isinstance(v, (int,float)):
            if v < 1 or v > 5:
                ok = False; break
    # Q41大面积缺失 (<2/5)
    q41_done = sum(1 for col in range(69,74) if ws.cell(row,col).value is not None and isinstance(ws.cell(row,col).value,(int,float)))
    if q41_done > 0 and q41_done < 2:
        ok = False
    # 逻辑矛盾
    rel = ws.cell(row, 31).value; sw = ws.cell(row, 20).value
    if rel is not None and sw is not None and isinstance(rel,(int,float)) and isinstance(sw,(int,float)):
        if rel >= 4 and sw <= 2:
            ok = False
    sched = ws.cell(row, 33).value; dist = ws.cell(row, 28).value
    if sched is not None and dist is not None and isinstance(sched,(int,float)) and isinstance(dist,(int,float)):
        if sched >= 4 and dist >= 5:
            ok = False
    # Straight-lining
    lik_vals = [ws.cell(row, col).value for col in likert_cols if ws.cell(row,col).value is not None and isinstance(ws.cell(row,col).value,(int,float))]
    if len(lik_vals) >= 6 and np.std(lik_vals) < 0.3 and len(set(lik_vals)) <= 2:
        ok = False

    if ok:
        valid.append(i)

NC = len(valid)
print(f"清洗后有效样本: N={NC}")
print()

# === 辅助函数 ===
def count_val(col, valid_rows, target_val):
    return sum(1 for i in valid_rows if ws.cell(i+2, col).value == target_val)

def count_all(col, valid_rows):
    return [ws.cell(i+2, col).value for i in valid_rows]

def mean_val(col, valid_rows):
    vals = [ws.cell(i+2, col).value for i in valid_rows if isinstance(ws.cell(i+2, col).value, (int,float))]
    return np.mean(vals) if vals else float('nan')

def pct(n, total):
    return f"{n}({n/total*100:.1f}%)"

# === 图表数据重新计算 ===
print("="*70)
print("各图表清洗后数据")
print("="*70)

# 年级分布 - check which column has grade
# Try col 8 or nearby
print("\n1. 年级分布 (doughnut)")
# checking columns around demographics... Let me check various columns
for col in range(1, 10):
    vals = [ws.cell(i+2, col).value for i in valid if ws.cell(i+2, col).value is not None]
    if vals:
        pass
# Let me just check col 8
grade_vals = [ws.cell(i+2, 8).value for i in valid if isinstance(ws.cell(i+2, 8).value, (int,float))]
if grade_vals:
    from collections import Counter
    gc = Counter(grade_vals)
    print(f"  年级分布: {dict(sorted(gc.items()))}")
    print(f"  总计数: {sum(gc.values())}")

# 性别
g = count_all(9, valid)
g1 = sum(1 for x in g if x==1)
g2 = sum(1 for x in g if x==2)
print(f"\n2. 性别: 男{g1}({g1/NC*100:.1f}%) 女{g2}({g2/NC*100:.1f}%)")

# 独生子女
oc = count_all(13, valid)
oc1 = sum(1 for x in oc if x==1)
oc2 = sum(1 for x in oc if x==2)
print(f"3. 独生子女: 是{oc1}({oc1/NC*100:.1f}%) 否{oc2}({oc2/NC*100:.1f}%)")

# 住校经历
board = count_all(15, valid)
b1 = sum(1 for x in board if x==1)
b2 = sum(1 for x in board if x==2)
print(f"4. 住校经历: 有{b1}({b1/NC*100:.1f}%) 无{b2}({b2/NC*100:.1f}%)")

# 舍友关系
rel = count_all(31, valid)
print(f"5. 舍友关系分布:")
for level in range(1,6):
    cnt = sum(1 for x in rel if x==level)
    labels = {1:'非常差',2:'比较差',3:'一般',4:'比较好',5:'非常好'}
    print(f"   {labels[level]}: {cnt}({cnt/len(rel)*100:.1f}%)")
# Chart data array
rel_data = [sum(1 for x in rel if x==lv) for lv in range(1,6)]
print(f"   Chart data: {rel_data}")

# 冲突解决方式
res = count_all(30, valid)
r1 = sum(1 for x in res if x==1)
r2 = sum(1 for x in res if x==2)
r4 = sum(1 for x in res if x==4)
print(f"6. 冲突解决: 直接{r1}({r1/NC*100:.1f}%) 第三方{r2}({r2/NC*100:.1f}%) 忍耐{r4}({r4/NC*100:.1f}%)")
print(f"   Chart data: [{r1},{r2},{r4}]")

# 调宿念头
sw = count_all(20, valid)
sw_data = [sum(1 for x in sw if x==lv) for lv in range(1,5)]
print(f"7. 调宿念头(1经常2偶尔3很少4从不): {sw_data}")

# 游戏频率
game = count_all(47, valid)
g1c = sum(1 for x in game if x==1)
g2c = sum(1 for x in game if x==2)
print(f"8. 游戏: 是{g1c}({g1c/NC*100:.1f}%) 否{g2c}({g2c/NC*100:.1f}%)")
print(f"   Chart data: [{g1c},{g2c}]")

# 锻炼频率
ex = count_all(61, valid)
ex_data = [sum(1 for x in ex if x==lv) for lv in range(1,6)]
print(f"9. 锻炼频率(1几乎不-5每天): {ex_data}")

# Q41排名均值
print(f"\n10. Q41因素重要度(分, 6-排名=越高越重要):")
q41_names = ['性格三观','生活习惯','兴趣爱好','经济条件','学习目标']
for idx, name in enumerate(q41_names):
    col = 69 + idx
    vals = [6 - ws.cell(i+2, col).value for i in valid if isinstance(ws.cell(i+2, col).value, (int,float))]
    if vals:
        print(f"   {name}: mean={np.mean(vals):.2f}")

# 焦虑
anx = [ws.cell(i+2, 23).value for i in valid if isinstance(ws.cell(i+2, 23).value, (int,float))]
print(f"\n11. 焦虑均值: {np.mean(anx):.2f}")

# 被吵醒
dist = [ws.cell(i+2, 28).value for i in valid if isinstance(ws.cell(i+2, 28).value, (int,float))]
print(f"12. 被吵醒均值: {np.mean(dist):.2f}")

# 冲突
conf = [ws.cell(i+2, 25).value for i in valid if isinstance(ws.cell(i+2, 25).value, (int,float))]
print(f"13. 冲突均值: {np.mean(conf):.2f}")

# 作息一致性
sched_vals = [ws.cell(i+2, 33).value for i in valid if isinstance(ws.cell(i+2, 33).value, (int,float))]
print(f"14. 作息一致性均值: {np.mean(sched_vals):.2f}")

# 硬件满意度
fac = [ws.cell(i+2, 37).value for i in valid if isinstance(ws.cell(i+2, 37).value, (int,float))]
print(f"15. 硬件满意度均值: {np.mean(fac):.2f}")

# 公平感
fair = [ws.cell(i+2, 44).value for i in valid if isinstance(ws.cell(i+2, 44).value, (int,float))]
print(f"16. 公平感均值(1公平3不公): {np.mean(fair):.2f}")

# 入睡时间
bed = [ws.cell(i+2, 45).value for i in valid if isinstance(ws.cell(i+2, 45).value, (int,float))]
print(f"17. 入睡时间(1早4晚): {np.mean(bed):.2f}")

# 睡眠噪音类型分布
sq = count_all(68, valid)
sq_dist = [sum(1 for x in sq if x==lv) for lv in range(1,8)]
print(f"18. 睡眠噪音类型(1-7): {sq_dist}")

# 生源地
org = count_all(10, valid)
o1 = sum(1 for x in org if x==1)
o2 = sum(1 for x in org if x==2)
print(f"19. 生源地: 城镇{o1}({o1/NC*100:.1f}%) 农村{o2}({o2/NC*100:.1f}%)")

# 改善建议 (需要找到对应列)
# 需要先确认列号
print(f"\n20. 独生子女冲突方式交叉表:")
# 独生×冲突解决
only_direct = sum(1 for i in valid if ws.cell(i+2,13).value==1 and ws.cell(i+2,30).value==1)
only_third = sum(1 for i in valid if ws.cell(i+2,13).value==1 and ws.cell(i+2,30).value==2)
only_endure = sum(1 for i in valid if ws.cell(i+2,13).value==1 and ws.cell(i+2,30).value==4)
non_direct = sum(1 for i in valid if ws.cell(i+2,13).value==2 and ws.cell(i+2,30).value==1)
non_third = sum(1 for i in valid if ws.cell(i+2,13).value==2 and ws.cell(i+2,30).value==2)
non_endure = sum(1 for i in valid if ws.cell(i+2,13).value==2 and ws.cell(i+2,30).value==4)
only_n = only_direct+only_third+only_endure
non_n = non_direct+non_third+non_endure
print(f"  独生(N={only_n}): 直接{only_direct}({only_direct/only_n*100:.0f}%) 第三方{only_third}({only_third/only_n*100:.0f}%) 忍耐{only_endure}({only_endure/only_n*100:.0f}%)")
print(f"  非独生(N={non_n}): 直接{non_direct}({non_direct/non_n*100:.0f}%) 第三方{non_third}({non_third/non_n*100:.0f}%) 忍耐{non_endure}({non_endure/non_n*100:.0f}%)")

print(f"\n{'='*70}")
print(f"SUMMARY: 所有百分比分母从110改为{NC}")
print(f"{'='*70}")
