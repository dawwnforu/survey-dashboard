# -*- coding: utf-8 -*-
"""宿舍人格量表效度交叉检验 + 其他因素统计分析"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 列映射（从col标题确认）:
#   8=年级, 9=性别(1男2女), 10=生源地, 13=独生子女(1是2否)
#   15=住校经历(1有2无), 20=调宿念头, 23=焦虑天数(0-7)
#   25=冲突次数, 28=被吵醒次数, 30=冲突解决(1直接2第三方4忍耐)
#   31=舍友关系(1-5), 32=卫生(1-5), 33=作息一致性(1-5)
#   37=硬件满意度, 44=公平感(1公平2不公平3非常不), 45=入睡时间范围(1-4)
#   47=游戏(1是2否), 61=户外锻炼, 68=睡眠质量(1-10)
#   69-73=分配因素重要性排名, 76=Q43人格类型(1-16)
#   79-83=改进方向(多选)

import openpyxl
import numpy as np
from scipy import stats as sp_stats

wb = openpyxl.load_workbook(r'D:\HuaweiMoveData\Users\xiaoy\Desktop\社会调查实践\问卷数据_110份.xlsx')
ws = wb.active

N = ws.max_row - 1
print(f"总样本数: {N}")

# === 读取所有数据 ===
def col_vals(col):
    """读取指定列的所有数值，跳过None"""
    vals = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row, col).value
        vals.append(v)
    return vals

# 基本人口学变量
gender     = col_vals(9)   # 1=男, 2=女
only_child = col_vals(13)  # 1=是, 2=否
boarding   = col_vals(15)  # 1=有, 2=无

# 核心连续/有序变量
anxiety     = col_vals(23)  # 焦虑天数 0-7
conflict    = col_vals(25)  # 月冲突次数
disturbed   = col_vals(28)  # 周被吵醒次数
resolve     = col_vals(30)  # 1=直接沟通,2=第三方,4=忍耐
relation    = col_vals(31)  # 舍友关系 1-5
hygiene     = col_vals(32)  # 卫生 1-5
schedule_ok = col_vals(33)  # 作息一致性 1-5
facility    = col_vals(37)  # 硬件满意度 1-5
fairness    = col_vals(44)  # 公平感(1公平2不公平3非常)
bedtime     = col_vals(45)  # 入睡范围 1=22前2=22-23 3=23-24 4=24后
gaming      = col_vals(47)  # 1=是(≥4次/周), 2=否
exercise    = col_vals(61)  # 户外锻炼频率
sleep_q     = col_vals(68)  # 睡眠质量 1-10
personality = col_vals(76)  # Q43人格 1-16
switch_idea = col_vals(20)  # 调宿念头

# 分配因素重要性 (排名, 数字越小越重要)
factor_life   = col_vals(70)  # 生活习惯契合
factor_char   = col_vals(69)  # 性格三观匹配
factor_hobby  = col_vals(71)  # 兴趣爱好相似
factor_econ   = col_vals(72)  # 经济条件相当
factor_study  = col_vals(73)  # 学习目标一致

# 特殊需求(多选, 1=勾选)
need_schedule = col_vals(63)  # 作息规律
need_quiet    = col_vals(64)  # 安静学习环境
need_allergy  = col_vals(65)  # 过敏/健康
need_religion = col_vals(66)  # 宗教/饮食
need_none     = col_vals(67)  # 无特殊需求

# 改进方向(多选)
improve_hw   = col_vals(79)  # 硬件设施
improve_mgmt = col_vals(80)  # 管理制度
improve_cult = col_vals(81)  # 文化建设
improve_svc  = col_vals(82)  # 后勤服务
improve_psy  = col_vals(83)  # 心理支持

def nanmean(arr):
    vals = [v for v in arr if v is not None and isinstance(v, (int, float))]
    return np.mean(vals) if vals else float('nan')

def nancount(arr):
    return sum(1 for v in arr if v is not None and isinstance(v, (int, float)))

def nanfilter(arr):
    return [v for v in arr if v is not None and isinstance(v, (int, float))]

# ============================================================
print("="*70)
print("=== 一、描述性统计 ===")
print("="*70)
print(f"性别: 男={sum(1 for g in gender if g==1)}, 女={sum(1 for g in gender if g==2)}")
print(f"独生子女: 是={sum(1 for o in only_child if o==1)}, 否={sum(1 for o in only_child if o==2)}")
print(f"住校经历: 有={sum(1 for b in boarding if b==1)}, 无={sum(1 for b in boarding if b==2)}")
print(f"焦虑天数均值: {nanmean(anxiety):.2f}")
print(f"被吵醒次数均值: {nanmean(disturbed):.2f}")
print(f"冲突次数均值: {nanmean(conflict):.2f}")
print(f"舍友关系均值: {nanmean(relation):.2f}")
print(f"卫生习惯均值: {nanmean(hygiene):.2f}")
print(f"睡眠质量均值: {nanmean(sleep_q):.2f}")

# ============================================================
# Q43 人格编码
# ============================================================
personality_map = {
    1:'ELTD',2:'ELTZ',3:'ELMD',4:'ELMZ',
    5:'EOTD',6:'EOTZ',7:'EOMD',8:'EOMZ',
    9:'ILTD',10:'ILTZ',11:'ILMD',12:'ILMZ',
    13:'IOTD',14:'IOTZ',15:'IOMD',16:'IOMZ'
}

# 找出Q43作答的索引
q43_indices = []
for i, p in enumerate(personality):
    if p is not None and isinstance(p, (int, float)) and 1 <= p <= 16:
        q43_indices.append(i)

n_q43 = len(q43_indices)
print(f"\nQ43人格作答数: {n_q43}")

# 拆分四维度
ei_vals = []  # E=1, I=0
lo_vals = []  # O=1(夜猫), L=0(早鸟)
tm_vals = []  # T=1(洁癖), M=0(随性)
dz_vals = []  # D=1(直球), Z=0(佛系)

for idx in q43_indices:
    code = personality_map[int(personality[idx])]
    ei_vals.append(1 if code[0]=='E' else 0)
    lo_vals.append(1 if code[1]=='O' else 0)
    tm_vals.append(1 if code[2]=='T' else 0)
    dz_vals.append(1 if code[3]=='D' else 0)

# 提取Q43作答者在其他变量上的值
def q43_subset(full_list):
    return [full_list[i] for i in q43_indices]

q43_anxiety   = q43_subset(anxiety)
q43_disturbed = q43_subset(disturbed)
q43_conflict  = q43_subset(conflict)
q43_relation  = q43_subset(relation)
q43_hygiene   = q43_subset(hygiene)
q43_fairness  = q43_subset(fairness)
q43_sched_ok  = q43_subset(schedule_ok)
q43_sleep_q   = q43_subset(sleep_q)
q43_bedtime   = q43_subset(bedtime)
q43_gaming    = q43_subset(gaming)
q43_resolve   = q43_subset(resolve)
q43_switch    = q43_subset(switch_idea)

# 派生变量
late_sleep = [1 if b is not None and isinstance(b,(int,float)) and b>=3 else 0 for b in q43_bedtime]  # 23点后
direct_comm = [1 if r==1 else 0 for r in q43_resolve]
gaming_bin = [1 if g==1 else (0 if g==2 else None) for g in q43_gaming]

# ============================================================
print("\n" + "="*70)
print(f"=== 二、Q43人格量表效度交叉检验 (N={n_q43}) ===")
print("="*70)

# --- Spearman ---
print("\n▶ Spearman 秩相关（双侧）")
dims = [
    ('E/I(社交能量)', ei_vals),
    ('L/O(作息节律)', lo_vals),
    ('T/M(空间秩序)', tm_vals),
    ('D/Z(冲突模式)', dz_vals)
]

var_pairs = [
    ('焦虑天数', q43_anxiety),
    ('被吵醒次数', q43_disturbed),
    ('冲突次数', q43_conflict),
    ('舍友关系', q43_relation),
    ('卫生习惯', q43_hygiene),
    ('公平感', q43_fairness),
    ('作息一致性', q43_sched_ok),
    ('睡眠质量', q43_sleep_q),
]

for dim_name, dim_vals in dims:
    for var_name, var_vals in var_pairs:
        # 对齐非None值
        pairs = [(d, v) for d, v in zip(dim_vals, var_vals) if v is not None and isinstance(v,(int,float))]
        if len(pairs) < 5:
            continue
        d_aligned = [p[0] for p in pairs]
        v_aligned = [p[1] for p in pairs]
        rho, p = sp_stats.spearmanr(d_aligned, v_aligned)
        sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
        if p < 0.10:
            print(f"  {dim_name} × {var_name}: rho={rho:.3f}{sig} (p={p:.4f}, n={len(pairs)})")

# 晚睡和直接沟通
for dim_name, dim_vals in dims:
    for aux_name, aux_vals in [('晚睡(23点后)', late_sleep), ('直接沟通', direct_comm)]:
        pairs = [(d, v) for d, v in zip(dim_vals, aux_vals) if v is not None]
        if len(pairs) < 5:
            continue
        d_aligned = [p[0] for p in pairs]
        v_aligned = [p[1] for p in pairs]
        rho, p = sp_stats.spearmanr(d_aligned, v_aligned)
        sig = '**' if p < 0.05 else ('†' if p < 0.10 else '')
        if p < 0.10:
            print(f"  {dim_name} × {aux_name}: rho={rho:.3f}{sig} (p={p:.4f}, n={len(pairs)})")

# --- 卡方 / Fisher ---
print("\n▶ 卡方 / Fisher 精确检验")
# D/Z × 沟通方式(直接 vs 非直接)
dz_direct_D = sum(1 for i in range(n_q43) if dz_vals[i]==1 and direct_comm[i]==1)
dz_avoid_D  = sum(1 for i in range(n_q43) if dz_vals[i]==1 and direct_comm[i]==0)
dz_direct_Z = sum(1 for i in range(n_q43) if dz_vals[i]==0 and direct_comm[i]==1)
dz_avoid_Z  = sum(1 for i in range(n_q43) if dz_vals[i]==0 and direct_comm[i]==0)
print(f"  D/Z × 沟通方式: D直球={dz_direct_D}直接/{dz_avoid_D}非直接, Z佛系={dz_direct_Z}直接/{dz_avoid_Z}非直接")
table = [[dz_direct_D, dz_avoid_D], [dz_direct_Z, dz_avoid_Z]]
try:
    chi2, p_chi, dof, exp = sp_stats.chi2_contingency(table)
    v_cramer = np.sqrt(chi2 / (n_q43 * 1))
    print(f"  χ²={chi2:.2f}, Cramér V={v_cramer:.3f}, p={p_chi:.4f}")
    odds, p_fish = sp_stats.fisher_exact(table)
    print(f"  Fisher exact: OR={odds:.2f}, p={p_fish:.4f}")
except Exception as e:
    print(f"  Error: {e}")

# L/O × 晚睡行为
lo_late_L = sum(1 for i in range(n_q43) if lo_vals[i]==0 and late_sleep[i]==1)
lo_notlate_L = sum(1 for i in range(n_q43) if lo_vals[i]==0 and late_sleep[i]==0)
lo_late_O = sum(1 for i in range(n_q43) if lo_vals[i]==1 and late_sleep[i]==1)
lo_notlate_O = sum(1 for i in range(n_q43) if lo_vals[i]==1 and late_sleep[i]==0)
print(f"  L/O × 晚睡: L早鸟={lo_late_L}晚睡/{lo_notlate_L}不晚睡, O夜猫={lo_late_O}晚睡/{lo_notlate_O}不晚睡")
table_lo = [[lo_late_L, lo_notlate_L], [lo_late_O, lo_notlate_O]]
try:
    chi2_lo, p_lo, _, _ = sp_stats.chi2_contingency(table_lo)
    v_lo = np.sqrt(chi2_lo / (n_q43 * 1))
    print(f"  χ²={chi2_lo:.2f}, Cramér V={v_lo:.3f}, p={p_lo:.4f}")
    odds_lo, p_fish_lo = sp_stats.fisher_exact(table_lo)
    print(f"  Fisher exact: OR={odds_lo:.2f}, p={p_fish_lo:.4f}")
except Exception as e:
    print(f"  Error: {e}")

# --- Mann-Whitney U ---
print("\n▶ Mann-Whitney U 检验")
# T/M × 卫生习惯
t_hyg = [q43_hygiene[i] for i in range(n_q43) if tm_vals[i]==1 and isinstance(q43_hygiene[i],(int,float))]
m_hyg = [q43_hygiene[i] for i in range(n_q43) if tm_vals[i]==0 and isinstance(q43_hygiene[i],(int,float))]
if len(t_hyg)>=3 and len(m_hyg)>=3:
    u_stat, p_mw = sp_stats.mannwhitneyu(t_hyg, m_hyg, alternative='two-sided')
    pooled = np.sqrt((np.std(t_hyg,ddof=1)**2+np.std(m_hyg,ddof=1)**2)/2)
    d_cohen = (np.mean(t_hyg)-np.mean(m_hyg))/pooled if pooled>0 else 0
    print(f"  T/M × 卫生习惯: U={u_stat:.0f}, p={p_mw:.4f}, Cohen d={d_cohen:.3f}")
    print(f"    T洁癖均值={np.mean(t_hyg):.2f}, M随性均值={np.mean(m_hyg):.2f}")

# T/M × 公平感
t_fair = [q43_fairness[i] for i in range(n_q43) if tm_vals[i]==1 and isinstance(q43_fairness[i],(int,float))]
m_fair = [q43_fairness[i] for i in range(n_q43) if tm_vals[i]==0 and isinstance(q43_fairness[i],(int,float))]
if len(t_fair)>=3 and len(m_fair)>=3:
    u2, p2 = sp_stats.mannwhitneyu(t_fair, m_fair, alternative='two-sided')
    print(f"  T/M × 公平感: U={u2:.0f}, p={p2:.4f}")

# --- 独立样本 t 检验 ---
print("\n▶ 独立样本 t 检验")
# D/Z × 舍友关系
d_rel = [q43_relation[i] for i in range(n_q43) if dz_vals[i]==1 and isinstance(q43_relation[i],(int,float))]
z_rel = [q43_relation[i] for i in range(n_q43) if dz_vals[i]==0 and isinstance(q43_relation[i],(int,float))]
if len(d_rel)>=3 and len(z_rel)>=3:
    t_dz, p_dz = sp_stats.ttest_ind(d_rel, z_rel)
    pooled_dz = np.sqrt((np.std(d_rel,ddof=1)**2+np.std(z_rel,ddof=1)**2)/2)
    d_dz = (np.mean(d_rel)-np.mean(z_rel))/pooled_dz if pooled_dz>0 else 0
    print(f"  D/Z × 舍友关系: t={t_dz:.2f}, p={p_dz:.4f}, Cohen d={d_dz:.3f}")
    print(f"    D直球均值={np.mean(d_rel):.2f}, Z佛系均值={np.mean(z_rel):.2f}")

# E/I × 公平感
e_fair = [q43_fairness[i] for i in range(n_q43) if ei_vals[i]==1 and isinstance(q43_fairness[i],(int,float))]
i_fair = [q43_fairness[i] for i in range(n_q43) if ei_vals[i]==0 and isinstance(q43_fairness[i],(int,float))]
if len(e_fair)>=3 and len(i_fair)>=3:
    t_ei, p_ei = sp_stats.ttest_ind(e_fair, i_fair)
    pooled_ei = np.sqrt((np.std(e_fair,ddof=1)**2+np.std(i_fair,ddof=1)**2)/2)
    d_ei = (np.mean(e_fair)-np.mean(i_fair))/pooled_ei if pooled_ei>0 else 0
    print(f"  E/I × 公平感: t={t_ei:.2f}, p={p_ei:.4f}, Cohen d={d_ei:.3f}")
    print(f"    E社牛均值={np.mean(e_fair):.2f}, I社恐均值={np.mean(i_fair):.2f}")

# E/I × 舍友关系
e_rel = [q43_relation[i] for i in range(n_q43) if ei_vals[i]==1 and isinstance(q43_relation[i],(int,float))]
i_rel = [q43_relation[i] for i in range(n_q43) if ei_vals[i]==0 and isinstance(q43_relation[i],(int,float))]
if len(e_rel)>=3 and len(i_rel)>=3:
    t_ei2, p_ei2 = sp_stats.ttest_ind(e_rel, i_rel)
    print(f"  E/I × 舍友关系: t={t_ei2:.2f}, p={p_ei2:.4f}")
    print(f"    E社牛均值={np.mean(e_rel):.2f}, I社恐均值={np.mean(i_rel):.2f}")

# === 效度排序 ===
print("\n▶ 四维度效度综合排序（定性）")
print("  基于效应量大小、p值显著性、方向一致性综合评估")
print("  具体排名请参考数据看板中的 validityChart")

# ============================================================
print("\n" + "="*70)
print(f"=== 三、其他因素交叉分析 (全样本 N={N}) ===")
print("="*70)

def ttest_groups(name_a, vals_a, name_b, vals_b):
    """打印两组对比 + t检验"""
    clean_a = nanfilter(vals_a)
    clean_b = nanfilter(vals_b)
    ma, mb = np.mean(clean_a), np.mean(clean_b)
    print(f"  {name_a}: 均值={ma:.2f}, N={len(clean_a)}")
    print(f"  {name_b}: 均值={mb:.2f}, N={len(clean_b)}")
    if len(clean_a)>=3 and len(clean_b)>=3:
        t, p = sp_stats.ttest_ind(clean_a, clean_b)
        pooled = np.sqrt((np.std(clean_a,ddof=1)**2+np.std(clean_b,ddof=1)**2)/2)
        d = (ma-mb)/pooled if pooled>0 else 0
        sig = '**' if p<0.05 else ('†' if p<0.10 else '')
        print(f"  t={t:.2f}, p={p:.4f}, Cohen d={d:.3f}{' '+sig if sig else ''}")
    print()

# 按性别分组
male_idx   = [i for i,g in enumerate(gender) if g==1]
female_idx = [i for i,g in enumerate(gender) if g==2]
game_idx   = [i for i,g in enumerate(gaming) if g==1]
nogame_idx = [i for i,g in enumerate(gaming) if g==2]
board_idx  = [i for i,b in enumerate(boarding) if b==1]
noboard_idx= [i for i,b in enumerate(boarding) if b==2]
only_idx   = [i for i,o in enumerate(only_child) if o==1]
notonly_idx= [i for i,o in enumerate(only_child) if o==2]
direct_idx = [i for i,r in enumerate(resolve) if r==1]
avoid_idx  = [i for i,r in enumerate(resolve) if r in [2,4]]

print("▶ 性别 × 焦虑天数")
ttest_groups('男', [anxiety[i] for i in male_idx], '女', [anxiety[i] for i in female_idx])

print("▶ 游戏 × 被吵醒次数")
ttest_groups('游戏者(≥4次/周)', [disturbed[i] for i in game_idx], '非游戏者', [disturbed[i] for i in nogame_idx])

print("▶ 住校经历 × 焦虑天数")
ttest_groups('有住校经历', [anxiety[i] for i in board_idx], '无住校经历', [anxiety[i] for i in noboard_idx])

print("▶ 独生子女 × 舍友关系")
ttest_groups('独生', [relation[i] for i in only_idx], '非独生', [relation[i] for i in notonly_idx])

print("▶ 沟通方式 × 舍友关系")
ttest_groups('直接沟通', [relation[i] for i in direct_idx], '第三方+忍耐', [relation[i] for i in avoid_idx])

print("▶ 游戏 × 焦虑天数")
ttest_groups('游戏者', [anxiety[i] for i in game_idx], '非游戏者', [anxiety[i] for i in nogame_idx])

print("▶ 住校经历 × 被吵醒次数")
ttest_groups('有住校经历', [disturbed[i] for i in board_idx], '无住校经历', [disturbed[i] for i in noboard_idx])

# 睡眠质量(Q40) × 焦虑 — 使用 Kruskal-Wallis（Q40为7类名义变量）
print("▶ 睡眠质量 × 焦虑天数 (Kruskal-Wallis, Q40为名义分类变量)")
sleep_cats = {}  # {category_value: [anxiety_scores]}
for i in range(N):
    sq = sleep_q[i]
    anx = anxiety[i]
    if isinstance(sq, (int,float)) and isinstance(anx, (int,float)) and 1 <= sq <= 7:
        cat = int(sq)
        if cat not in sleep_cats:
            sleep_cats[cat] = []
        sleep_cats[cat].append(anx)
for cat in sorted(sleep_cats.keys()):
    vals = sleep_cats[cat]
    print(f"  类别{cat} ({len(vals)}人): 焦虑均值={np.mean(vals):.2f}")
groups = [sleep_cats[c] for c in sorted(sleep_cats.keys()) if len(sleep_cats[c]) >= 3]
if len(groups) >= 3:
    h_stat, p_kw = sp_stats.kruskal(*groups)
    print(f"  Kruskal-Wallis H={h_stat:.2f}, p={p_kw:.4f}")

# 作息一致性 × 被吵醒 (Spearman)
print("\n▶ 作息一致性 × 被吵醒次数 (Spearman)")
sc_pairs = [(schedule_ok[i], disturbed[i]) for i in range(N) if isinstance(schedule_ok[i],(int,float)) and isinstance(disturbed[i],(int,float))]
if len(sc_pairs) >= 5:
    sx = [p[0] for p in sc_pairs]
    sy = [p[1] for p in sc_pairs]
    rho, p = sp_stats.spearmanr(sx, sy)
    print(f"  rho={rho:.3f}, p={p:.4f}, N={len(sc_pairs)}")

# 卫生习惯 × 舍友关系 (Spearman)
print("\n▶ 卫生习惯 × 舍友关系 (Spearman)")
hy_pairs = [(hygiene[i], relation[i]) for i in range(N) if isinstance(hygiene[i],(int,float)) and isinstance(relation[i],(int,float))]
if len(hy_pairs) >= 5:
    sx = [p[0] for p in hy_pairs]
    sy = [p[1] for p in hy_pairs]
    rho, p = sp_stats.spearmanr(sx, sy)
    print(f"  rho={rho:.3f}, p={p:.4f}, N={len(hy_pairs)}")

# 分配因素重要性
print("\n▶ 分配因素重要性排名（原始均分，越低越重要）")
factors = [
    ('生活习惯契合', factor_life),
    ('性格三观匹配', factor_char),
    ('兴趣爱好相似', factor_hobby),
    ('学习目标一致', factor_study),
    ('经济条件相当', factor_econ),
]
for name, vals in factors:
    clean = nanfilter(vals)
    if clean:
        print(f"  {name}: {np.mean(clean):.2f}")

# 特殊需求覆盖率
print("\n▶ 特殊需求覆盖率 (%)")
needs_list = [
    ('作息规律匹配', need_schedule),
    ('安静学习环境', need_quiet),
    ('过敏/健康', need_allergy),
    ('宗教/饮食', need_religion),
    ('无特殊需求', need_none),
]
for name, vals in needs_list:
    yes = sum(1 for v in vals if v==1)
    total = sum(1 for v in vals if v is not None)
    if total > 0:
        print(f"  {name}: {yes}/{total} = {yes/total*100:.1f}%")

print("\n" + "="*70)
print("分析完成。所有数值均为从 Excel 实际计算的结果。")
print("="*70)
